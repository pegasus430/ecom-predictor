import os
import io
import zipfile
import traceback
import boto
import uuid
import requests
from boto.s3.key import Key
from datetime import datetime
from lxml import etree
from urlparse import urljoin

import xlrd
import xlwt
import openpyxl
from xlutils.save import save


def cvt_to_xls(src_file_path, dst_file_path):
    if src_file_path.endswith('.xls'):
        # resave file to make it readable
        src_book = xlrd.open_workbook(src_file_path)
        save(src_book, dst_file_path)
    elif src_file_path.endswith('.xlsx'):
        src_book = openpyxl.load_workbook(src_file_path, read_only=True)
        dst_book = xlwt.Workbook()

        for sheet_name in src_book.sheetnames:
            src_sheet = src_book[sheet_name]
            dst_sheet = dst_book.add_sheet(sheet_name)

            for row_index, row in enumerate(src_sheet.rows):
                for cell_index, cell in enumerate(row):
                    dst_sheet.write(row_index, cell_index, cell.value)

        dst_book.save(dst_file_path)


def generate_new_file_path(file_path, report_name, report_date):
    file_name = file_path.split('/')[-1]
    file_dir = '/'.join(file_path.split('/')[:-1])

    user = file_name.split('_')[0]

    report_name = report_name.replace(' ', '_')

    new_file_name = '{0}_{1}_{2}.xls'.format(user, report_name, report_date)
    new_file_path = file_dir + '/' + new_file_name

    return new_file_path


class WalmartRetailCrawlerException(Exception):
    pass


class WalmartRetailCrawler():
    LOGIN_URL = "https://rllogin.wal-mart.com/rl_security/rl_logon.aspx"

    resources_dir = '/tmp'
    bucket_name = 'retail-link-images'

    def __init__(self, proxy=None, proxy_type=None, user_agent=None):

        self.test_mode = False
        self._responses = []

        self.session = requests.Session()

        if proxy:
            proxy = '{}://{}'.format(proxy_type or 'http', proxy)

            self.session.proxies.update({
                'http': proxy,
                'https': proxy,
            })

        self.session.headers.update({
            'User-Agent': user_agent or 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:32.0) Gecko/20100101 Firefox/32.0'
        })

    def get_test_report(self, name_report):
        test_reports_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'test_reports')
        reports = os.listdir(test_reports_dir)

        for report in reports:
            if name_report == os.path.splitext(report)[0]:
                src_file_path = os.path.join(test_reports_dir, report)
                dst_file_name = generate_new_file_path('{}/user_{}'.format(self.resources_dir, report),
                                                       name_report,
                                                       datetime.now().strftime('%Y-%m-%d'))
                cvt_to_xls(src_file_path, dst_file_name)

                return dst_file_name

        raise WalmartRetailCrawlerException('Test report not found')

    def get_report(self, name_report):
        try:
            report_file = self._get_report(name_report)

            report_name = os.path.splitext(os.path.basename(report_file))[0]

            responses_url = self._upload_responses(report_name)
            print 'Responses: {}'.format(responses_url)
        except Exception as e:
            responses_url = self._upload_responses(uuid.uuid4())
            print 'Responses: {}'.format(responses_url)

            setattr(e, 'responses', responses_url)
            raise

        return report_file

    def _get_report(self, name_report):
        if self.test_mode:
            return self.get_test_report(name_report)

        # load reports via API
        response = self.session.get('https://retaillink.wal-mart.com/rl_home_services/api/Site/GetRequestList')
        self.save_response(response.content, 'API response')

        if response.status_code != requests.codes.ok:
            raise WalmartRetailCrawlerException('API error: {}'.format(response.status_code))

        for req in response.json():
            if req.get('ReqName') == name_report:
                if isinstance(name_report, unicode):
                    name_report = name_report.encode('utf-8')

                if req.get('StatusCode') not in ('D', 'R'):
                    raise WalmartRetailCrawlerException('Report "{}" not ready. Current status: {}'.format(
                        name_report,
                        req.get('StatusDesc')))

                report_filename = req.get('RetFileName', '')
                report_filename_ext = req.get('RetFileExt', '')

                src_file_path = "{}/{}.{}".format(self.resources_dir,
                                                  report_filename.rsplit('/', 1)[-1], report_filename_ext)

                if not os.path.exists(src_file_path):
                    report_url = urljoin('https://retaillink.wal-mart.com',
                                         '{}.{}'.format(report_filename, report_filename_ext))
                    report_res = self.session.get(report_url)

                    if report_res.status_code != requests.codes.ok:
                        self.save_response(report_res.content, 'Report response')
                        raise WalmartRetailCrawlerException('Report download error: {}'.format(report_res.status_code))

                    with open(src_file_path, 'wb') as report_file:
                        report_file.write(report_res.content)

                dst_file_name = generate_new_file_path(src_file_path, name_report, req.get('RunDate', '').split(' ')[0])
                cvt_to_xls(src_file_path, dst_file_name)

                return dst_file_name

        raise WalmartRetailCrawlerException('Report not found')

    def do_login(self, user, password):
        try:
            self._do_login(user, password)
        except Exception as e:
            responses_url = self._upload_responses(uuid.uuid4())
            print 'Responses: {}'.format(responses_url)

            setattr(e, 'responses', responses_url)
            raise

    def _do_login(self, user, password):
        if user == 'user' and password == 'pass':
            self.test_mode = True
            return

        response = self.session.get(self.LOGIN_URL)
        self.save_response(response.content, 'Login page')

        if response.status_code != requests.codes.ok:
            raise WalmartRetailCrawlerException('Login page is inaccessible: {}'.format(response.status_code))

        tree = etree.HTML(response.content)

        view_state = tree.xpath(".//input[@id='__VIEWSTATE']/@value")
        event_validation = tree.xpath(".//input[@id='__EVENTVALIDATION']/@value")

        if not view_state or not event_validation:
            raise WalmartRetailCrawlerException('Invalid login form')

        auth_data = {
            'txtUser': user,
            'txtPass': password,
            '__VIEWSTATE': view_state[0],
            '__EVENTVALIDATION': event_validation[0],
            'Login': 'Logon'
        }

        response = self.session.post(self.LOGIN_URL, data=auth_data)
        self.save_response(response.content, 'Auth result')

        if response.status_code != requests.codes.ok:
            raise WalmartRetailCrawlerException('Auth error: {}'.format(response.status_code))

        if 'changePassword.aspx' in response.content:
            raise WalmartRetailCrawlerException('Password Expired')

        if 'rl_logon.aspx' in response.content:
            raise WalmartRetailCrawlerException('Login Failed')

        if 'rl_home/' not in response.content:
            raise WalmartRetailCrawlerException('Unknown auth error')

    def save_response(self, content, description=''):
        timestamp = datetime.now().strftime('%Y-%m-%d__%H_%M_%S_%f')
        response_path = os.path.join('{}/{}.txt'.format(self.resources_dir, timestamp))

        try:
            with open(response_path, 'w') as response_file:
                response_file.write(content)
                print 'Response: {} ({})'.format(description, response_path)
                self._responses.append(response_path)
        except:
            print 'ERROR: Could not save response:  {} ({})'.format(description, response_path)

    def _upload_responses(self, report_name):
        if self._responses:
            buf = io.BytesIO()

            with zipfile.ZipFile(buf, 'w', compression=zipfile.ZIP_DEFLATED) as zip_file:
                for response in self._responses:
                    if os.path.exists(response):
                        zip_file.write(response, os.path.basename(response))
                    else:
                        print 'WARNING: Response {} does not exist'.format(response)

            return self._upload_content_to_s3('{}_responses.zip'.format(report_name), buf.getvalue())

    def _upload_content_to_s3(self, key, content):
        if self.bucket_name:
            try:
                s3_conn = boto.connect_s3()
                s3_bucket = s3_conn.get_bucket(self.bucket_name, validate=False)

                s3_key = Key(s3_bucket)
                s3_key.key = key
                s3_key.set_metadata("Content-Type", 'application/octet-stream')
                s3_key.set_contents_from_string(content)

                url = s3_key.generate_url(expires_in=0, query_auth=False)
                return url.split('?')[0]
            except:
                print 'ERROR: Could not upload content {} to S3: {}'.format(key, traceback.format_exc())
        else:
            print 'Bucket property is empty. Skip uploading {} to S3'.format(key)

if __name__ == '__main__':
    password = "pass"
    user = "user"

    crawler = WalmartRetailCrawler()
    crawler.do_login(user, password)

    path = crawler.get_report("Dashboard Report - Retail Sales")
    print path
