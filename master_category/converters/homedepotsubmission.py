import json
import uuid
import urllib2
import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from xlsxwriter.workbook import Workbook

try:
    import cStringIO as StringIO
except ImportError:
    import StringIO


class HomeDepotProxy(object):
    output_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    output_ext = '.xlsx'

    prd_submission_url = 'http://submissions.contentanalyticsinc.com/api/v1/submissions'
    test_submission_url = 'http://submissions.contentanalyticsinc.com:8080/api/v1/submissions'

    def __init__(self, input_file, channel):
        self._input_file = input_file
        self.products = []
        name = os.path.splitext(os.path.basename(input_file))[0]
        self.category = name.split('_')[-1]
        self._channel = channel
        if self._channel == 'home-depot':
            self.submission_url = self.prd_submission_url
        elif self._channel == 'homedepot-thd-qa':
            self.submission_url = self.test_submission_url
        else:
            self.submission_url = self.test_submission_url  # default

    @staticmethod
    def _to_str(s):
        if s is None:
            return u''
        return unicode(s)

    def _load_products(self):
        wb = load_workbook(filename=self._input_file, read_only=True)
        sheet_name = wb.get_sheet_names()[0]
        sheet = wb[sheet_name]

        data_rows_idx_start = 7
        all_rows = list(sheet.rows)
        required_row = all_rows[2]
        ids_row = all_rows[0]
        rows_count = len(all_rows)
        columns_count = len(ids_row)

        for data_idx in range(data_rows_idx_start, rows_count + 1):
            product = {}
            i = 1  # skip very first column
            while i < columns_count:
                next_i = i + 1 if i < columns_count - 1 else None
                column = get_column_letter(i+1) + str(data_idx)  # letters starts from 1(A)
                next_column = get_column_letter(next_i+1) + str(data_idx) if next_i is not None else None
                two_in_one = False
                is_required = required_row[i].value and required_row[i].value.lower().strip() == 'required'
                # determine if columns have one identificator
                if next_column and ids_row[next_i].value is None and ids_row[i].value is not None:
                    two_in_one = True

                attr_id = ids_row[i].value
                attr_val = self._to_str(sheet[column].value)
                if two_in_one:
                    attr_val += u' ' + self._to_str(sheet[next_column].value)
                    attr_val = attr_val.strip()
                if not attr_val:
                    if is_required:
                        # TODO raise exception if require field empty?
                        pass
                else:
                    product[attr_id] = attr_val
                i += 1
                if two_in_one:
                    i += 1
            product['__CATEGORY'] = self.category
            product['__NAME'] = product.get('0d96a759-6ab3-424c-9b73-3aec0d09d2f6', 'No name provided')
            self.products.append(product)

    def _send_submission(self):
        options = {
            'products': self.products,
            'channel': self._channel
        }
        submission_data = {
            'submission': {
                'type': 'user_uploaded',
                'retailer': 'homedepot.com',
                'options': options
            }
        }

        feed_id = uuid.uuid4().get_hex()

        submission_request = urllib2.Request(
            self.submission_url,
            data=json.dumps(submission_data),
            headers={
                'X-API-KEY': 'alo4yu8fj30ltb3r',
                'X-FEED-ID': feed_id,
                'Content-Type': 'application/json'
            })

        response = json.loads(urllib2.urlopen(submission_request).read())

        if response.get('status') == 'error':
            return False, response
        return True, response

    def convert(self):
        self._load_products()
        result, response = self._send_submission()
        output = StringIO.StringIO()
        book = Workbook(output)
        sheet = book.add_worksheet('products')
        for i, row in enumerate(self.products):
                sheet.write(i, 0, json.dumps(row))

        result_sheet = book.add_worksheet('result')
        result_sheet.write(0, 0, 'Submission posted')
        result_sheet.write(1, 0, result)
        result_sheet.write(0, 1, 'Response')
        result_sheet.write(1, 1, json.dumps(response))

        book.close()
        # construct response
        output.seek(0)

        yield output.read()
