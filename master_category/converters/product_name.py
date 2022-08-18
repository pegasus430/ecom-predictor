import csv
import re
from xlsxwriter.workbook import Workbook

try:
    import cStringIO as StringIO
except ImportError:
    import StringIO

from openpyxl import load_workbook


class Echo(object):
    @staticmethod
    def write(value):
        return value


class ProductNameConverter(object):
    csv = 1
    xlsx = 2

    def __init__(self, input_file):
        self.input_file_type = 0
        self.output_type = None
        self.products = self._read_input_file(input_file)

    def _read_input_file(self, input_file):
        # Trying to read as xlsx firstly and if it's fail try read as csv
        try:
            wb = load_workbook(filename=input_file, read_only=True)
            self.input_file_type = self.xlsx
            self.output_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            # iterate all sheets
            first_sheet_name = wb.get_sheet_names()[0]
            first_sheet = wb[first_sheet_name]
            for row in first_sheet.iter_rows():
                product_name = row[0].value or ''
                yield product_name
        except:
            try:
                with open(input_file, 'rU') as f:
                    input_csv = csv.reader(f)
                    self.input_file_type = self.csv
                    self.output_type = 'text/csv'
                    for row in input_csv:
                        yield row[0]
            except:
                raise Exception("Can't read provided file. Supported files are CSV and XLSX")

    def convert(self):
        results = [['Input', 'Output', 'Warnings']]
        for product in self.products:
            converted, comments = self._convert(product)
            results.append([product, converted, comments])
        if self.input_file_type == self.csv:
            buf = Echo()
            output_csv = csv.writer(buf)
            for row in results:
                yield output_csv.writerow(row)
            return
        if self.input_file_type == self.xlsx:
            output = StringIO.StringIO()
            book = Workbook(output)
            sheet = book.add_worksheet('results')
            for i, row in enumerate(results):
                for j, cell in enumerate(row):
                    sheet.write(i, j, cell)
            book.close()
            # construct response
            output.seek(0)

            yield output.read()

    def _convert(self, product):
        raise NotImplementedError


class WalmartGrocery(ProductNameConverter):
    def _convert(self, product):
        result = product
        try:
            result, comment = self._process_conversion(product)
        except Exception as e:
            return result, e.message
        return result, comment

    @staticmethod
    def _get_comments(product):
        # if product contain slash we need to mention that
        comments = []
        if '/' in product:
            comments.append('contain "/"')
        return '; '.join(comments)

    def _process_conversion(self, product):
        product = self._to_lower(product)
        product = self._to_abbreviations(product)
        product = self._split(product)
        product = self._replace(product)
        product = self._add_commas(product)
        product = self._camel_case(product)
        product = self._remove_trademarks(product)
        comment = self._get_comments(product)
        return product, comment

    @staticmethod
    def _to_lower(product):
        words = [
            'ounce',
            'ounces',
            'oz',
            'count',
            'ct',
            'pack',
            'pk',
            'fluid',
            'fl',
            'fld',
            'milliliter',
            'millilitre',
            'ml'
        ]
        for word in words:
            product = re.sub('(?<![a-zA-Z]){}(?![a-zA-Z])'.format(word), word, product, flags=re.I)
        return product

    @staticmethod
    def _to_abbreviations(product):
        words = {
            'oz': ['ounces', 'ounce'],
            'ct': ['count'],
            'pk': ['pack'],
            'fl': ['fluid', 'fld'],
            'ml': ['milliliter', 'millilitre']
        }
        for abbr, word_list in words.iteritems():
            for word in word_list:
                product = re.sub('(?<![a-zA-Z]){}(?![a-zA-Z])'.format(word), abbr, product)
        return product

    @staticmethod
    def _split(product):
        words = ['oz', 'ct', 'pk', 'fl', 'ml']
        for word in words:
            product = re.sub('(\d){}(?![a-zA-Z])'.format(word), '\g<1> {}'.format(word), product)
            product = re.sub('(?<![a-zA-Z]){}\.'.format(word), word, product)
        return product

    @staticmethod
    def _replace(product):
        # remove "pk" if it's don't have number before
        product_words = product.split()
        start = 0
        while 'pk' in product_words[start:]:
            pos = product_words.index('pk', start)
            if pos < 1 or not product_words[pos - 1].isdigit():
                product_words = product_words[:pos] + product_words[pos + 1:]
                continue
            start = pos + 1
        product = ' '.join(product_words)

        if re.search('(?<![a-zA-Z])ct(?![a-zA-Z])', product) and re.search('(?<![a-zA-Z])pk(?![a-zA-Z])', product):
            raise Exception('Has ct and pk')
        product = re.sub('(?<![a-zA-Z])pk(?![a-zA-Z])', 'ct', product)
        return product

    @staticmethod
    def _add_commas(product):
        # each number (if it's quantity/count/amount) should come after comma
        measure_words = ['oz', 'ct', 'pk', 'fl', 'ml']

        def _need_comma(_word, _prev, _next):
            if not _word.replace('.', '', 1).isdigit():
                return False
            if _prev and _prev.endswith(','):
                return False
            if not _next or _next.lower() not in measure_words:
                return False
            return True

        product_words = product.split()
        updated_words = []
        for i, word in enumerate(product_words):
            next = product_words[i + 1] if i + 1 < len(product_words) else None
            prev = product_words[i - 1] if i > 1 else None
            if updated_words and _need_comma(word, prev, next):
                updated_words[-1] += ','
            updated_words.append(word)
        product = ' '.join(updated_words)
        return product

    @staticmethod
    def _camel_case(product):
        def _need_convert(word):
            if word.strip(',.)(') in exceptions:
                return False
            return True
        exceptions = ['XL', 'XXL', 'XXXL', 'XS']
        exceptions += ['oz', 'ct', 'pk', 'fl', 'ml']
        # if there a word with all upper chars -> replace it with CamelCase
        product = ' '.join(word.title() if _need_convert(word) else word for word in product.split())
        return product

    @staticmethod
    def _remove_trademarks(product):
        trademarks = ['(TM)', '(SM)', '(R)', '(C)']
        res_product = []
        for word in product.split():
            if word.upper().strip(',.') in trademarks:
                for splitter in '.,':
                    if word.endswith(splitter) and res_product and not res_product[-1].endswith(splitter):
                        res_product[-1] += splitter
                continue
            for trademark in trademarks:
                if trademark in word.upper():
                    word = word.replace(trademark, '')
            res_product.append(word)
        product = ' '.join(res_product)
        return product


class WupcUpc(ProductNameConverter):
    UPC_TO_WUPC = 1
    WUPC_TO_UPC = 2
    WUPC_INTEGERS_PART_LENGTH = 11
    UPC_INTEGERS_PART_LENGTH = 12
    WUPC_PATTERN = '[0]{2}([0-9]{11})([\w]*)'
    UPC_PATTERN = '([0-9]{12})([\w]*)'

    def __init__(self, input_file, conversion_type=UPC_TO_WUPC):
        super(WupcUpc, self).__init__(input_file)
        self._conversion_type = conversion_type

    @staticmethod
    def _wupc_to_upc(name):
        # check format
        pattern = re.compile(WupcUpc.WUPC_PATTERN)
        match = pattern.match(name)
        if not match:
            raise Exception('Wrong wUPC format (expected "{}")'.format(WupcUpc.WUPC_PATTERN))
        # split numeric part and left part
        numeric_part, ending_part = match.groups()
        # compute sum of odd and even elements
        odd = sum([int(ch) for ch in numeric_part[::2]])
        even = sum([int(ch) for ch in numeric_part[1::2]])
        # multiple odd by 3
        odd *= 3
        result = odd + even
        # remainder of division by 10
        remainder = result % 10
        # check digit is 10 - remainder
        check_digit = 10 - remainder
        return numeric_part + str(check_digit) + ending_part

    @staticmethod
    def _upc_to_wupc(name):
        # check format
        pattern = re.compile(WupcUpc.UPC_PATTERN)
        match = pattern.match(name)
        if not match:
            raise Exception('Wrong UPC format (expected "{}")'.format(WupcUpc.UPC_PATTERN))
        # split numeric part and left part
        numeric_part, ending_part = match.groups()
        # add leading zeros
        numeric_part = '00' + numeric_part
        # remove final (check) digit
        numeric_part = numeric_part[:-1]
        return numeric_part + ending_part

    def _convert(self, name):
        result = name
        try:
            result, comment = self._process_conversion(name)
        except Exception as e:
            return result, e.message
        return result, comment

    def _process_conversion(self, name):
        if self._conversion_type == WupcUpc.WUPC_TO_UPC:
            return self._wupc_to_upc(name), ''
        elif self._conversion_type == WupcUpc.UPC_TO_WUPC:
            return self._upc_to_wupc(name), ''
        else:
            raise Exception('Bad conversion type')
