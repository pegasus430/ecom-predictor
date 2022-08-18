import ast
import csv
import time

import os
import re
import boto
import openpyxl
import requests
import xlrd
import xlwt
from boto.s3.key import Key
from flask import Flask, render_template, render_template_string, request, send_from_directory, jsonify
from openpyxl import load_workbook
from werkzeug import secure_filename
from xlrd import open_workbook
from xlutils.copy import copy
from datetime import datetime

from xlsxpatch import XlsxPatch


# Initialize the Flask application
app = Flask(__name__)

# This is the path to the upload directory
app.config['UPLOAD_DIR'] = '/var/tmp/mc_exports/upload/'
app.config['EXPORT_DIR'] = '/var/tmp/mc_exports/export/'
app.config['TEMPLATE_DIR'] = '/var/tmp/mc_exports/templates/'

if not os.path.exists(app.config['UPLOAD_DIR']):
    os.makedirs(app.config['UPLOAD_DIR'])

if not os.path.exists(app.config['EXPORT_DIR']):
    os.makedirs(app.config['EXPORT_DIR'])

if not os.path.exists(app.config['TEMPLATE_DIR']):
    os.makedirs(app.config['TEMPLATE_DIR'])


# This is the path to the template file
app.config['TEMPLATE_MAPPING_FILE'] = '/var/tmp/mc_exports/Template Mapping File.xlsx'
app.config['MANDATORY_MAPPING_FIELDS'] = {'upc', 'gtin', 'tool_id', 'asin', 'tcin',
                                          'vendor_code_id', 'vendor_item_sku_number', 'mpn'}

# These are the extension that we are accepting to be uploaded
app.config['ALLOWED_EXTENSIONS'] = {'csv'}

app.config['BUCKET_NAME'] = 'retailer-template-files'

app.config['MC_API_URL'] = 'http://{0}.contentanalyticsinc.com/api/products?api_key={1}'
app.config['MC_UNMATCHED_API_URL'] = 'http://{0}.contentanalyticsinc.com/api/products/unmatched?api_key={1}'

TEMPLATE_STRING = '''
<h3>Predefined template</h3>
<form name=mc_export action={{url_for('mc_export')}} enctype=multipart/form-data method=post>
<input type=file name=file></input><br/>
<br/>
<input type=text name='retailer', placeholder=Retailer></input><br/>
<input type=text name='server', placeholder=Server></input><br/>
<br/>
<input type="radio" name="updated_content" value="true">updated_content = true<br/>
<input type="radio" name="updated_content" value="false" checked>updated_content = false<br/>
<br/>
<button type=submit>Export</button>
</form>

<h3>Custom template</h3>
<form name=mc_export action={{url_for('mc_export')}} enctype=multipart/form-data method=post>
<input type=file name=file></input><br/>
<br/>
<input type=text name='retailer', placeholder="Retailer (optional)"></input><br/>
<input type=text name='server', placeholder=Server></input><br/>
<input type=text name='file_name' placeholder="File name"></input><br/>
<input type=text name='file_type' placeholder="File type (optional)"></input><br/>
<input type=hidden name='field[]' value="product_name"></input>
<input type=hidden name='field[]' value="price"></input>
<input type=hidden name='field[]' value="currency"></input>
<input type=hidden name='field_name[product_name]' value="Product Name"></input>
<input type=hidden name='field_name[price]' value="Price"></input>
<input type=hidden name='field_name[currency]' value="Currency"></input>
<br/>
<input type="radio" name="updated_content" value="true">updated_content = true<br/>
<input type="radio" name="updated_content" value="false" checked>updated_content = false<br/>
<br/>
<button type=submit>Export</button>
</form>
'''


class Error(Exception):
    def __init__(self, message, status_code=200):
        Exception.__init__(self)
        self.message = {
            'error': True,
            'message': message
            }
        self.status_code = status_code


@app.errorhandler(Error)
def handle_error(error):
    response = jsonify(error.message)
    response.status_code = error.status_code
    return response


# For a given file, return whether it's an allowed type or not
def allowed_file(filename):
    return '.' in filename and \
        filename.rsplit('.', 1)[1] in app.config['ALLOWED_EXTENSIONS']


_mc_api_keys = {}


def get_mc_api_key(server):
    if server not in _mc_api_keys:
        print 'Requesting API key for server {}'.format(server)
        api_url = 'https://{server}.contentanalyticsinc.com/api/token?' \
                  'username=api@cai-api.com&password=jEua6jLQFRjq8Eja'.format(server=server)
        response = requests.get(api_url)
        response.raise_for_status()
        data = response.json()
        _mc_api_keys[server] = data['api_key']
    return _mc_api_keys[server]


def parse_input_file(file_path, server):
    ids = {}
    results = []

    # Either get info from MC API or use the info in the file
    with open(file_path, 'r') as f:

        # request in groups of 1000
        i = 0
        for row in csv.DictReader(f.read().splitlines()):
            k = i / 100

            if k not in ids:
                ids[k] = []

            ids[k].append(row.get('CAID'))
            i += 1

        for k in ids:
            data = {'filter[products][]': ids[k],
                    'product[apply_product_changes]': True}

            try:
                r = requests.get(app.config['MC_API_URL'].format(server, get_mc_api_key(server)),
                                 params=data).json()
            except requests.exceptions.ConnectionError:
                raise Error('Couldn\'t connect to host: {0}'.format(server))
            except Exception as e:
                print 'Error requesting MC data: %s' % e
                raise Error('Error requesting MC data from host: {0}'.format(server))

            for p in r.get('products', []):
                if p.get('image_urls'):
                    p['images'] = p['image_urls']
                elif p.get('images') and not isinstance(p['images'], list):
                    p['images'] = ast.literal_eval(re.sub("u?'", '', p['images']))
                results.append(p)

    return results


def unmatched_input_data(input_data, server):
    unmatched_products = []

    request_limit = 1000

    try:
        for i in range(0, len(input_data), request_limit):
            products = input_data[i:i + request_limit]

            data = {'filter[products][]': map(lambda x: x['id'], products)}

            unmatched_data = requests.get(app.config['MC_UNMATCHED_API_URL'].format(server, get_mc_api_key(server)),
                                          params=data).json()

            if unmatched_data and isinstance(unmatched_data, list):
                for unmatched_product in unmatched_data:
                    product = next((p for p in products if p.get('url') == unmatched_product.get('url')), None)

                    if product:
                        unmatched_product.update({k: v for k, v in product.iteritems()
                                                  if k in app.config['MANDATORY_MAPPING_FIELDS']})

                        if 'new_images' in unmatched_product:
                            unmatched_product['images'] = product.get('image_urls')
                            unmatched_product.pop('new_images')

                        unmatched_products.append(unmatched_product)
    except Exception as e:
        print 'ERROR: {}'.format(e)

    return unmatched_products


def clean(v):
    if isinstance(v, basestring):
        return v.strip()
    return v


def parse_template_mapping_file(file_path):
    template_mapping_info_json = {"column_name_list": [], "map_info_by_retailer": {}}
    workbook = xlrd.open_workbook(file_path)
    worksheet = workbook.sheet_by_index(0)

    # Change this depending on how many header rows are present
    # Set to 0 if you want to include the header data.
    offset = 0

    for i, row in enumerate(range(worksheet.nrows)):
        if i <= offset:  # (Optionally) skip headers
            for j, col in enumerate(range(worksheet.ncols)):
                template_mapping_info_json["column_name_list"].append(clean(worksheet.cell_value(i, j)))
            continue

        r = {}

        for j, col in enumerate(range(worksheet.ncols)):
            r[clean(worksheet.cell_value(0, j))] = clean(worksheet.cell_value(i, j))

        template_mapping_info_json["map_info_by_retailer"][clean(worksheet.cell_value(i, 1)).lower()] = r

    return template_mapping_info_json


def create_template(template_name, worksheet_name, headers):
    if template_name.lower().endswith('.xls'):
        wb = xlwt.Workbook()
        ws = wb.add_sheet(worksheet_name)
        for col_index, header in enumerate(headers):
            ws.write(0, col_index, header)
        wb.save(app.config['TEMPLATE_DIR'] + template_name)
    elif template_name.lower().endswith('.csv'):
        wb = open(app.config['TEMPLATE_DIR'] + template_name, 'wb')
        ws = csv.writer(wb)
        ws.writerow(headers)
        wb.close()
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = worksheet_name
        for col_index, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_index, value=header)
        wb.save(app.config['TEMPLATE_DIR'] + template_name)
    return template_name


def fill_out_template(file_name, save_file_name, input_data_info, template_mapping_info, retailer):
    worksheet_name = template_mapping_info["map_info_by_retailer"][retailer]['Tab']
    print 'WS', worksheet_name

    use_xlsxpatch = False
    if retailer == 'kroger_com':
        use_xlsxpatch = True

    is_xls = False
    is_csv = False

    if file_name.lower().endswith('.xls'):
        rb = open_workbook(app.config['TEMPLATE_DIR'] + file_name)
        rs = rb.sheet_by_name(worksheet_name)
        wb = copy(rb)
        ws = wb.get_sheet(worksheet_name)
        is_xls = True
    elif file_name.lower().endswith('.csv'):
        rb = open(app.config['TEMPLATE_DIR'] + file_name)
        rs = csv.reader(rb)
        wb = open(app.config['EXPORT_DIR'] + save_file_name, 'wb')
        ws = csv.writer(wb)
        is_csv = True
    else:
        rb = load_workbook(filename=app.config['TEMPLATE_DIR'] + file_name)
        rs = rb.get_sheet_by_name(worksheet_name)
        if use_xlsxpatch:
            wb = XlsxPatch(app.config['TEMPLATE_DIR'] + file_name)
            ws = wb.get_sheet_by_name(worksheet_name)
        else:
            wb = rb
            ws = rs

    COLUMN_HEADER_ROW = int(template_mapping_info["map_info_by_retailer"][retailer]['Column Header Row'])

    template_file_headers = {}

    if is_xls:
        for col_index in range(0, rs.ncols): 
            header = clean(rs.cell_value(COLUMN_HEADER_ROW-1, col_index))
            if header:
                template_file_headers[header] = col_index
    elif is_csv:
        for _ in range(COLUMN_HEADER_ROW):
            row = rs.next()
            ws.writerow(row)

        headers = map(lambda x: x.strip(), row)

        for col_index, header in enumerate(headers):
            template_file_headers[header] = col_index
    else:
        for col_index in range(1, rs.max_column+1):
            header = clean(rs.cell(row=COLUMN_HEADER_ROW, column=col_index).value)
            if header:
                template_file_headers[header] = col_index

    CONTENT_START_ROW = int(template_mapping_info["map_info_by_retailer"][retailer]['Content Start Row'])

    if is_csv:
        for _ in range(CONTENT_START_ROW-COLUMN_HEADER_ROW-1):
            row = rs.next()
            ws.writerow(row)

    def _write_val(_row_index, _mapped_col, _val):
        _val = _val.encode('utf-8') if isinstance(_val, unicode) else _val
        if is_xls:
            ws.write(_row_index + CONTENT_START_ROW - 1, _mapped_col, _val)
        elif is_csv:
            row_values[_mapped_col] = _val
        elif use_xlsxpatch:
            ws.write(_row_index + CONTENT_START_ROW, _mapped_col, _val)
        else:
            if isinstance(_val, str):
                _val = re.sub(openpyxl.cell.cell.ILLEGAL_CHARACTERS_RE, '', _val)
            ws.cell(row=_row_index + CONTENT_START_ROW, column=_mapped_col).value = _val

    for row_index, row in enumerate(input_data_info):
        row_values = [''] * len(template_file_headers.keys())

        if retailer == 'kroger_com':
            if row.get('gtin'):
                row['upc'] = format_upc(row['gtin'])
            if not row.get('attributes'):
                row['attributes'] = {}
            if 'common' not in row['attributes']:
                row['attributes']['common'] = {}
            if 'ImageLocation' not in row['attributes']['common']:
                row['attributes']['common']['ImageLocation'] = 'Supplier has image(s)'
            diff = row.get('diff', {})
            row['change product_name'] = 'No' if diff.get('diff_product_name', 'f') == 'f' else 'Yes'
            row['change brand'] = 'No'  # not tracked
            row['change itemsize'] = 'No'  # not tracked
            row['change image'] = 'No' if diff.get('diff_images', '0') == '0' else 'Yes'

        if retailer == 'jet' and not row.get('bullets') and row.get('long_description'):
            row['bullets'] = bullets_from_description(row.get('long_description'))

        if retailer.startswith('dollargeneral'):
            row['long_description'] = extend_long_description(row)

        if row.get('bullets'):
            row['bullets_all'] = join_bullets(row['bullets'])

        row['current_date'] = datetime.now().strftime('%Y-%m-%d')

        for column in row:
            if column in ['images', 'bullets']:
                for i in range(20):
                    new_column = column + ' ' + str(i+1)

                    mapped_column_name = template_mapping_info["map_info_by_retailer"][retailer].get(new_column)

                    if mapped_column_name:
                        mapped_col = template_file_headers.get(mapped_column_name)

                        if mapped_col is not None and row.get(column) and len(row[column]) > i:
                            _write_val(row_index, mapped_col, row[column][i])
            elif column == 'attributes' and row[column]:
                attributes = row[column]
                for attr_type, certain_attributes in attributes.iteritems():
                    for attr_name, attr_val in certain_attributes.iteritems():
                        combined_column_name = ' '.join([attr_type, 'attribute', attr_name]).lower()
                        mapped_column_name = (
                            template_mapping_info["map_info_by_retailer"][retailer].get(combined_column_name)
                        )
                        if mapped_column_name:
                            mapped_col = template_file_headers.get(mapped_column_name)
                            if mapped_col is not None:
                                _write_val(row_index, mapped_col, attr_val)
            else:
                mapped_column_name = template_mapping_info["map_info_by_retailer"][retailer].get(column)

                if mapped_column_name:
                    mapped_col = template_file_headers.get(mapped_column_name)

                    if mapped_col is not None:
                        value = row[column]

                        if value and mapped_column_name == 'Bulleted Copy':
                            bullets = bullets_from_description(value)
                            if bullets:
                                value = description_from_bullets(bullets)

                        _write_val(row_index, mapped_col, value)

        if is_csv:
            ws.writerow(row_values)
    if is_csv:
        rb.close()
        wb.close()
    else:
        if use_xlsxpatch:
            rb.close()
        wb.save(app.config['EXPORT_DIR'] + save_file_name)

    return save_file_name


def format_upc(upc):
    r = list(reversed(upc))

    check_digit = int(r[0])
    sum_digits = r[1:]

    by3 = [3 * int(x) for i, x in enumerate(sum_digits) if i % 2 == 0]
    by1 = [int(x) for i, x in enumerate(sum_digits) if i % 2 == 1]

    summed_total = sum(by3 + by1) + check_digit

    if summed_total % 10 == 0 and len(upc.lstrip('0')) > 10:
        # remove check digit
        upc = upc[:-1]

    return upc[-13:].zfill(13)


def join_bullets(bullets):
    bullets_all = u'<ul>'

    for bullet in bullets:
        bullets_all += u'<li>{}</li>'.format(bullet)

    bullets_all += u'</ul>'

    return bullets_all


def bullets_from_description(description):
    match = re.search(r'<ul>(.*?)</ul>', description)

    return re.findall(r'(?:<li>(.*?)</li>)+?', match.group(1) if match else description)


def description_from_bullets(bullets):
    description = u''

    for bullet in bullets:
        description += u'{} {}\n'.format(unichr(8226), bullet)

    return description


def extend_long_description(product):
    long_description = product.get('long_description')

    if not long_description:
        long_description = u''
    else:
        long_description = u'<p>{}</p>'.format(long_description)

    bullets = product.get('bullets')
    if bullets:
        long_description += u'<p><strong>Key Features</strong>:</p><ul>' + \
                            u''.join(map(lambda x: u'<li>' + x + u'</li>', bullets)) + \
                            u'</ul>'

    directions = product.get('usage_directions')
    if directions:
        long_description += u'<p><strong>Directions</strong>: {}</p>'.format(directions)

    ingredients = product.get('ingredients')
    if ingredients:
        long_description += u'<p><strong>Ingredients</strong>: {}</p>'.format(ingredients)

    allergens = product.get('caution_warnings_allergens')
    if allergens:
        long_description += u'<p><strong>Allergens<strong>: {}</p>'.format(allergens)

    return long_description


# This route will show a form to perform an AJAX request
# jQuery is loaded to execute the request and update the
# value of the operation
@app.route('/')
def index():
    return render_template('index.html')


def fetch_from_s3(retailer=None):
    s3_conn = boto.connect_s3(is_secure=False)
    s3_bucket = s3_conn.get_bucket(app.config['BUCKET_NAME'], validate=False)

    found = None
    for key in s3_bucket.list():
        if retailer:
            if key.name.startswith(retailer):
                with open(app.config['TEMPLATE_DIR'] + key.name, 'w') as f:
                    f.write(key.read())
                found = key
                break
        else:
            if key.name == 'Template Mapping File.xlsx':
                with open(app.config['TEMPLATE_MAPPING_FILE'], 'w') as f:
                    f.write(key.read())
                found = key
                break

    if found:
        print 'Got {0} from S3'.format(found.name)
        return found.name

    raise Error('Failed to locate template file for {0}'.format(retailer))


# Route that will process the file upload
@app.route('/mc_export', methods=['GET', 'POST'])
def mc_export():
    print 'RECEIVED REQUEST', request

    if request.method == 'GET':
        return render_template_string(TEMPLATE_STRING)

    f = request.files.get('file') or request.form.get('file')
    print 'F', f

    retailer = request.args.get('retailer') or request.form.get('retailer')
    print 'R', retailer
    server = request.args.get('server') or request.form.get('server')
    print 'S', server
    updated_content = request.args.get('updated_content') or request.form.get('updated_content')
    print 'U', updated_content

    file_name = request.args.get('file_name') or request.form.get('file_name')
    print 'N', file_name
    file_type = request.args.get('file_type') or request.form.get('file_type') or ''
    print 'T', file_type
    worksheet = request.args.get('worksheet') or request.form.get('worksheet') or 'Custom'
    print 'W', worksheet
    fields = request.args.getlist('field[]') or request.form.getlist('field[]')
    print 'F', fields

    if not f:
        raise Error('Missing input file', 400)

    if not allowed_file(f.filename):
        raise Error('Incorrect file type: the input file should be csv format', 400)

    if file_name:
        if not retailer:
            retailer = 'custom'

        if not fields:
            index = 0
            while True:
                arg_name = 'field[%s]' % index
                field = request.args.get(arg_name) or request.form.get(arg_name)
                if not field:
                    break
                fields.append(field)
                index += 1
            print 'F', fields

        if not fields:
            raise Error('Missing param: field[] or field[0]', 400)

        headers = []
        template_mapping_info = {
            "map_info_by_retailer": {
                retailer: {
                    'Tab': worksheet,
                    'Column Header Row': 1,
                    'Content Start Row': 2,
                }
            }
        }
        for field in fields:
            arg_name = 'field_name[%s]' % field
            field_name = request.args.get(arg_name) or request.form.get(arg_name) or field
            print arg_name, '=', field_name
            headers.append(field_name)
            template_mapping_info["map_info_by_retailer"][retailer][field] = field_name

        export_file_name = template_name = str(int(time.time())) + '_' + file_name
        if file_type:
            extension = '.' + file_type.lower()
            if not template_name.lower().endswith(extension):
                template_name += extension

        create_template(template_name, worksheet, headers)
        print app.config['TEMPLATE_DIR'] + template_name
    else:
        if not retailer:
            raise Error('Missing param: retailer', 400)

        retailer = retailer.strip().lower()

        # fetch the template mapping file from S3
        fetch_from_s3()

        template_mapping_info = parse_template_mapping_file(app.config["TEMPLATE_MAPPING_FILE"])

        if retailer not in template_mapping_info["map_info_by_retailer"]:
            raise Error('Retailer {0} doesn\'t exist in template mapping file'.format(retailer))

        template_name = fetch_from_s3(retailer=retailer)
        print app.config['TEMPLATE_DIR'] + template_name

        export_file_name = str(int(time.time())) + '_' + template_name

    # Make the filename safe, remove unsupported chars
    filename = secure_filename(f.filename)

    # Move the file from the temporal folder to
    # the upload folder we setup
    f.save(os.path.join(app.config['UPLOAD_DIR'], filename))

    input_data_info = parse_input_file(os.path.join(app.config['UPLOAD_DIR'], filename), server)
    # IDI is populated
    if updated_content == 'true':
        input_data_info = unmatched_input_data(input_data_info, server)

    fill_out_template(template_name, export_file_name, input_data_info, template_mapping_info, retailer)

    # Upload the file to S3
    try:
        conn = boto.connect_s3(is_secure=False)
        bucket = conn.get_bucket(app.config['BUCKET_NAME'], validate=False)

        current_timestamp = int(export_file_name.split('_', 1)[0])
        current_file_name = export_file_name.split('_', 1)[-1]

        # remove any files with the same name in the exports directory which are over a day old
        for k in bucket.list('exports/'):
            filename_parts = k.name.split('/')[-1].split('_', 1)
            if filename_parts[-1] == current_file_name:
                timestamp = int(filename_parts[0])
                age_in_days = (current_timestamp - timestamp) / (60*60*24)
                if age_in_days >= 1:
                    bucket.delete_key(k)

        k = Key(bucket)
        k.key = 'exports/' + export_file_name

        f = open(app.config['EXPORT_DIR'] + export_file_name, 'r')
        k.set_contents_from_file(f)
        f.close()

    except Exception as e:
        print 'Error uploading file: %s' % e
        raise Error('Error uploading file')

    return jsonify({
        'error': False,
        'file': 'https://s3.amazonaws.com/{0}/exports/{1}'.format(
            app.config['BUCKET_NAME'], export_file_name)
        })


# This route is expecting a parameter containing the name
# of a file. Then it will locate that file on the upload
# directory and show it on the browser, so if the user uploads
# an image, that image is going to be show after the upload
@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_DIR'],
                               filename)

if __name__ == '__main__':
    app.run(
        host="0.0.0.0",
        port=int("8999"),
        debug=True
    )
