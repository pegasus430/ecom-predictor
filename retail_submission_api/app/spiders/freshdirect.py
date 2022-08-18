import traceback
import boto
import os

from openpyxl import load_workbook
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart

from . import SubmissionSpider, SubmissionSpiderError


class FreshDirectSubmissionSpider(SubmissionSpider):
    retailer = 'freshdirect.com'
    driver_engine = None  # don't use web driver

    template_file = './freshdirect/Fresh Direct - NIF.xlsx'

    def _template_name(self):
        name = 'Fresh Direct - NIF {date}.xlsx'.format(
            date=datetime.now().strftime('%Y%m%d%H%M%S')
        )

        return name

    def _template_generate(self, products):
        wb = load_workbook(self.template_file)

        worksheet = wb.active

        row_index = 3

        for product in products:
            upc = product.get('upc')
            if upc:
                worksheet.cell(column=3, row=row_index, value=upc)

            product_name = product.get('product_name')
            if product_name:
                worksheet.cell(column=4, row=row_index, value=product_name)

            common_attributes = (product.get('attributes') or {}).get('common') or {}

            base_unit_of_measure = common_attributes.get('BaseUnitOfMeasure')
            if base_unit_of_measure:
                worksheet.cell(column=5, row=row_index, value=base_unit_of_measure)

            net_weight = common_attributes.get('NetWeight')
            if net_weight:
                worksheet.cell(column=8, row=row_index, value=net_weight)

            weight_unit = common_attributes.get('WeightUnit')
            if weight_unit:
                worksheet.cell(column=9, row=row_index, value=weight_unit)

            product_length = common_attributes.get('ProductLength')
            if product_length:
                worksheet.cell(column=10, row=row_index, value=product_length)

            product_width = common_attributes.get('ProductWidth')
            if product_width:
                worksheet.cell(column=11, row=row_index, value=product_width)

            product_height = common_attributes.get('ProductHeight')
            if product_height:
                worksheet.cell(column=12, row=row_index, value=product_height)

            order_unit = common_attributes.get('OrderUnit')
            if order_unit:
                worksheet.cell(column=13, row=row_index, value=order_unit)

            units_per_case = common_attributes.get('UnitsPerCase')
            if units_per_case:
                worksheet.cell(column=14, row=row_index, value=units_per_case)

            net_unit_cost = common_attributes.get('NetUnitCost')
            if net_unit_cost:
                worksheet.cell(column=15, row=row_index, value=net_unit_cost)

            unit_msrp = common_attributes.get('UnitMSRP')
            if unit_msrp:
                worksheet.cell(column=16, row=row_index, value=unit_msrp)

            case_upc = common_attributes.get('CaseUPC')
            if case_upc:
                worksheet.cell(column=17, row=row_index, value=case_upc)

            case_length = common_attributes.get('CaseLength')
            if case_length:
                worksheet.cell(column=18, row=row_index, value=case_length)

            case_width = common_attributes.get('CaseWidth')
            if case_width:
                worksheet.cell(column=19, row=row_index, value=case_width)

            case_height = common_attributes.get('CaseHeight')
            if case_height:
                worksheet.cell(column=20, row=row_index, value=case_height)

            gross_case_cost = common_attributes.get('GrossCaseCost')
            if gross_case_cost:
                worksheet.cell(column=21, row=row_index, value=gross_case_cost)

            case_oi_discount = common_attributes.get('CaseOIDiscount')
            if case_oi_discount:
                worksheet.cell(column=22, row=row_index, value=case_oi_discount)

            net_fd_case_cost = common_attributes.get('NetFDCaseCost')
            if net_fd_case_cost:
                worksheet.cell(column=23, row=row_index, value=net_fd_case_cost)

            lead_time_in_days = common_attributes.get('LeadTimeInDays')
            if lead_time_in_days:
                worksheet.cell(column=25, row=row_index, value=lead_time_in_days)

            bottle_deposit = common_attributes.get('BottleDeposit')
            if bottle_deposit:
                worksheet.cell(column=26, row=row_index, value=bottle_deposit)

            juice_percent = common_attributes.get('JuicePercent')
            if juice_percent:
                worksheet.cell(column=27, row=row_index, value=juice_percent)

            link_to_product_web_page = common_attributes.get('LinkToProductWebPage')
            if link_to_product_web_page:
                worksheet.cell(column=28, row=row_index, value=link_to_product_web_page)

            attributes_1 = common_attributes.get('Attributes1')
            if attributes_1:
                worksheet.cell(column=30, row=row_index, value=attributes_1)

            attributes_2 = common_attributes.get('Attributes2')
            if attributes_2:
                worksheet.cell(column=31, row=row_index, value=attributes_2)

            attributes_3 = common_attributes.get('Attributes3')
            if attributes_3:
                worksheet.cell(column=32, row=row_index, value=attributes_3)

            ti_hi_information_cases = common_attributes.get('TiHiInformationCases')
            if ti_hi_information_cases:
                worksheet.cell(column=33, row=row_index, value=ti_hi_information_cases)

            ti_hi_information_layers = common_attributes.get('TiHiInformationLayers')
            if ti_hi_information_layers:
                worksheet.cell(column=34, row=row_index, value=ti_hi_information_layers)

            min_delivered_shelf_life = common_attributes.get('MinDeliveredShelfLife')
            if min_delivered_shelf_life:
                worksheet.cell(column=36, row=row_index, value=min_delivered_shelf_life)

            upc_length = common_attributes.get('UPCLength')
            if upc_length:
                worksheet.cell(column=37, row=row_index, value=upc_length)

            case_upc_length = common_attributes.get('CaseUPCLength')
            if case_upc_length:
                worksheet.cell(column=38, row=row_index, value=case_upc_length)

            description = product.get('description')
            if description:
                worksheet.cell(column=39, row=row_index, value=description)

            row_index += 1

        template_filename = self.get_file_path_for_result(self._template_name())

        wb.save(template_filename)

        return template_filename

    def _email_send(self, subject, body, to, attachment_filename):
        msg = MIMEMultipart()

        msg['Subject'] = subject
        msg['From'] = 'noreply@contentanalyticsinc.com'
        msg['To'] = to

        msg.attach(MIMEText(body))

        attachment = MIMEApplication(open(attachment_filename, 'rb').read())
        attachment.add_header('Content-Disposition', 'attachment', filename=os.path.split(attachment_filename)[-1])
        msg.attach(attachment)

        try:
            ses = boto.connect_ses()

            ses.send_raw_email(msg.as_string(), source=msg['From'], destinations=[msg['To']])

            return True
        except:
            self.logger.error('Can not send email: {}'.format(traceback.format_exc()))
            return False

    def task_text(self, options, products, server, **kwargs):
        missing_options = {'email'} - set(options.keys())
        if missing_options:
            raise SubmissionSpiderError('Missing options: {}'.format(', '.join(missing_options)))

        self.logger.info('Template generating')
        template_filename = self._template_generate(products)

        self.logger.info('Template sending')

        subject = 'Fresh Direct Content Submission - {date}'.format(
            date=datetime.now().strftime('%Y-%m-%d')
        )

        body = 'Please find the content updates in the attached file.'

        if not self.sandbox and options.get('do_submit'):
            result = self._email_send(
                subject, body, options.get('email'), template_filename)
        else:
            subject = 'TEST SUBMISSION. ' + subject

            result = self._email_send(
                subject, body, 'support@contentanalyticsinc.com', template_filename)

        if not result:
            raise SubmissionSpiderError('Email sending error')

        self.logger.info('Email was sent')

    def task_images(self, options, products, server, criteria, **kwargs):
        missing_options = {'email'} - set(options.keys())
        if missing_options:
            raise SubmissionSpiderError('Missing options: {}'.format(', '.join(missing_options)))

        images = self._export_media(criteria, server)

        self.logger.info('Sending images: {}'.format(images))

        subject = 'Fresh Direct Image Submission - {date}'.format(
            date=datetime.now().strftime('%Y-%m-%d')
        )

        body = 'Please find the image updates in the attached file.'

        if not self.sandbox and options.get('do_submit'):
            result = self._email_send(
                subject, body, options.get('email'), images)
        else:
            subject = 'TEST SUBMISSION. ' + subject

            result = self._email_send(
                subject, body, 'support@contentanalyticsinc.com', images)

        if not result:
            raise SubmissionSpiderError('Email sending error')

        self.logger.info('Email was sent')
