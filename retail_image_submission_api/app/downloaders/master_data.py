import shutil
import traceback
import urllib2
import urlparse

from cloudinary.uploader import upload
from openpyxl import Workbook
from PIL import Image

from . import ImageSubmissionDownloader, ImageSubmissionDownloaderError


class MasterDataImageSubmissionDownloader(ImageSubmissionDownloader):
    retailer = 'master data'

    def _cloudinary(self, options, products):
        self.logger.info('[START UPLOAD IMAGES]')

        urls = options.get('urls') or []
        if products:
            for product in products:
                product_id = product.get('id')

                image_urls = product.get('image_urls')

                if not image_urls:
                    self.logger.warn('Product {} has not images'.format(product_id))
                    continue

                new_images = product.get('new_images') or []

                self.logger.debug('Collecting images for product {}'.format(product_id))

                for image_index, image_url in enumerate(image_urls):
                    if options.get('differences_only') and image_url not in new_images:
                        continue

                    urls.append(image_url)
        elif not urls:
            self.logger.warn('No images to process')

        image_type = options.get('image_type')
        image_square = options.get('image_square')
        image_resize = options.get('image_resize')
        image_min_side_dimension = options.get('image_min_side_dimension')
        image_max_side_dimension = options.get('image_max_side_dimension')
        transformation = []
        if image_resize and image_max_side_dimension:
            transformation.append('c_limit,w_{max},h_{max}')
        if image_square:
            transformation.append('c_pad,ar_1,b_white')
        if transformation and image_min_side_dimension:
            condition = 'if_w_gte_{min}_or_h_gte_{min}'
            if len(transformation) == 1:
                transformation[0] = ','.join((condition, transformation[0]))
            else:
                transformation = [condition] + transformation + ['if_end']
        eager = '/'.join(transformation).format(min=image_min_side_dimension,
                                                max=image_max_side_dimension)

        result_urls = []
        for url in urls:
            self.logger.debug('Uploading image: {}'.format(url))

            public_id = None
            parts = urlparse.urlparse(url)
            if parts.netloc == 'productspace.pepsico.com':
                args = urlparse.parse_qs(parts.query)
                public_id = args.get('fn', args.get('fileName', [None]))[-1]
                if public_id:
                    public_id = public_id.rsplit('.', 1)[0]
                    self.logger.debug('Public id: {}'.format(public_id))

            result_url = ''
            try:
                image = upload(url, eager=eager, format=image_type, public_id=public_id)
                if eager:
                    result = image['eager'][0]
                else:
                    result = image
            except:
                self.logger.error('Can not upload image {}: {}'.format(url, traceback.format_exc()))
            else:
                if 'secure_url' in result:
                    result_url = result['secure_url']
                else:
                    self.logger.error('Can not process image {}: {}'.format(url, result.get('reason')))
            result_urls.append(result_url)

        self.logger.info('[END UPLOAD IMAGES]')
        return urls, result_urls

    def task_cloudinary(self, options, products):
        try:
            urls, result_urls = self._cloudinary(options, products)

            self.logger.info('[START DOWNLOAD IMAGES]')

            for image_url in result_urls:
                if not image_url:
                    continue
                try:
                    self.logger.debug('Loading image: {}'.format(image_url))
                    res = urllib2.urlopen(image_url)
                except:
                    self.logger.error('Can not load image {}: {}'.format(image_url, traceback.format_exc()))
                else:
                    if res.getcode() != 200:
                        self.logger.error('Can not load image {}: response code {}, content: {}'.format(
                            image_url, res.getcode(), res.read()))
                    else:
                        with open(self.get_file_path_for_result(image_url.rsplit('/', 1)[-1]), 'wb') as image_file:
                            shutil.copyfileobj(res, image_file)

            self.logger.info('[END DOWNLOAD IMAGES]')
        except:
            self.logger.error('Submission error: {}'.format(traceback.format_exc()))
            raise ImageSubmissionDownloaderError('Submission failed')

    def task_cloudinary_urls(self, options, products):
        try:
            urls, result_urls = self._cloudinary(options, products)
            wb = Workbook()
            ws = wb.active
            for index, (old_url, new_url) in enumerate(zip(urls, result_urls), 1):
                ws.cell(row=index, column=1, value=old_url)
                ws.cell(row=index, column=2, value=new_url)
            wb.save(self.get_file_path_for_result('images.xlsx'))
        except:
            self.logger.error('Submission error: {}'.format(traceback.format_exc()))
            raise ImageSubmissionDownloaderError('Submission failed')

    def task_images(self, options, products):
        try:
            self.logger.info('[START DOWNLOAD IMAGES]')

            for product in products:
                product_id = product.get('id')

                upc = product.get('upc')

                if not upc:
                    self.logger.error('Product {} has not UPC'.format(product_id))
                    continue

                image_urls = product.get('image_urls')

                if not image_urls:
                    self.logger.warn('Product {} has not images'.format(product_id))
                    continue

                self.logger.debug('Loading images for product {}'.format(product_id))

                for image_index, image_url in enumerate(image_urls):
                    try:
                        self.logger.debug('Loading image: {}'.format(image_url))
                        res = urllib2.urlopen(image_url, timeout=60)
                    except:
                        self.logger.error('Can not load image {}: {}'.format(image_url, traceback.format_exc()))
                    else:
                        if res.getcode() != 200:
                            self.logger.error('Can not load image {}: response code {}, content: {}'.format(
                                image_url, res.getcode(), res.read()))
                        else:
                            image = Image.open(res)

                            if not image.mode.startswith('RGB'):
                                image = image.convert('RGB')

                            if options.get('image_resize'):
                                width, height = image.size
                                dimension = options.get('image_min_side_dimension')

                                if dimension and width != height and width >= dimension and height >= dimension:
                                    ratio = float(dimension) / max(width, height)

                                    width = int(width * ratio)
                                    height = int(height * ratio)

                                    image = image.resize((width, height), Image.LANCZOS)

                                    new_image = Image.new('RGB', (dimension, dimension), (255, 255, 255))
                                    new_image.paste(image, ((dimension - width) / 2, (dimension - height) / 2),
                                                    image if image.mode == 'RGBA' else None)

                                    image = new_image

                            image_type = options.get('image_type', 'jpg')
                            image_name = '{}.{}'.format(self._get_image_name(upc, image_index), image_type)

                            image.save(self.get_file_path_for_result(image_name))

            self.logger.info('[END DOWNLOAD IMAGES]')
        except:
            self.logger.error('Submission error: {}'.format(traceback.format_exc()))
            raise ImageSubmissionDownloaderError('Submission failed')

    def _get_image_name(self, upc, index):
        return '{name}_{suffix:02}'.format(name=upc, suffix=index)
