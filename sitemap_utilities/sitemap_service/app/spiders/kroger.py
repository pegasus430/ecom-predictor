import csv
from datetime import datetime
from HTMLParser import HTMLParser
import json
import re
import requests
import uuid
import urlparse
import traceback

from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import PatternFill
from openpyxl.worksheet.write_only import WriteOnlyCell

from . import SitemapSpider, SitemapSpiderError


class KrogerSitemapSpider(SitemapSpider):
    retailer = 'kroger.com'

    headers = {
        'Content-Type': 'application/json;charset=utf-8',
        'User-Agent': 'Slackbot-LinkExpanding 1.0 (+https://api.slack.com/robots)',
    }

    STORES_SITEMAP_URL = 'https://www.kroger.com/storelocator-sitemap.xml'

    def __init__(self, *args, **kwargs):
        super(KrogerSitemapSpider, self).__init__(*args, **kwargs)

        self.stores_cache = {}

    def task_geo_report(self, options):
        self._check_options(options, params=['urls', 'stores'])

        report_name = options.get('request_name') or 'geo_report'

        with open(self.get_file_path_for_result('{}.csv'.format(report_name)), 'wb') as geo_report_file:
            csv_writer = csv.writer(geo_report_file)
            csv_writer.writerow(['Zip Code', 'Store ID', 'Product Name', 'URL', 'Availability', 'Price'])

            stores = options.get('stores', [])

            failed_urls = []
            for store in stores:
                zip_code, division_id, store_id = self._parse_store(store)

                if not division_id or not store_id:
                    self.logger.warn('Cannot get store for zip code {} and store id {}'.format(zip_code, store_id))
                    continue

                self.logger.info('Loading info for division id {} and store id {}'.format(division_id, store_id))

                for url in options.get('urls', []):
                    try:
                        self.logger.debug('Checking store {}/{}: {}'.format(division_id, store_id, url))

                        product_id = urlparse.urlparse(url).path.split('/')[-1]
                        product_info = self._get_products_info([product_id], division_id, store_id)
                        if product_info:
                            product_info = product_info[0]
                        else:
                            product_info = {}

                        csv_writer.writerow([zip_code, store_id,
                                             product_info.get('name'),
                                             url,
                                             product_info.get('availability', False),
                                             product_info.get('price')])
                    except:
                        self.logger.error('Cannot process url: {}'.format(traceback.format_exc()))
                        if url not in failed_urls:
                            failed_urls.append(url)
            if failed_urls:
                self.save_failed_urls(failed_urls)
                raise SitemapSpiderError('Some urls cannot be processed')

    def _parse_store(self, store):
        zip_code = None
        store_id = None

        if isinstance(store, dict):
            zip_code = store.get('zip_code')
            store_id = store.get('store_id')
        elif isinstance(store, (list, tuple)):
            if len(store) == 1:
                zip_code = store[0],
            elif len(store) > 1:
                zip_code = store[0],
                store_id = store[1]
        else:
            zip_code = store

        division_id, store_id = self._get_store_id(zip_code, store_id)

        return zip_code, division_id, store_id

    def _get_store_id(self, zip_code, store_id=None):
        if not store_id:
            store_id = None

        if zip_code not in self.stores_cache:
            self.logger.info('Search store for zip code: {} and store id {}'.format(zip_code, store_id))

            store_search_url = 'https://www.kroger.com/stores/api/graphql'

            store_search_params = {
                'query': 'query modalStoreSearch($searchText: String!) {modalStoreSearch(searchText: $searchText) {stores '
                         '{divisionNumber vanityName storeNumber phoneNumber showShopThisStoreAndPreferredStoreButtons '
                         'distance address {addressLine1 addressLine2 city countryCode stateCode zip} formattedHours '
                         '{displayName displayHours isToday} fulfillmentMethods {hasPickup hasDelivery} departments {code}}'
                         ' eligibleDeliveryStore {zipCode storeNumber divisionNumber}}}',
                'operationName': 'modalStoreSearch',
                'variables': {'searchText': zip_code},
            }

            response = requests.post(store_search_url, json=store_search_params, headers=self.headers, timeout=60)

            if self._check_response(response):
                data = response.json()

                stores = data.get('data', {}).get('modalStoreSearch', {}).get('stores', [])

                if stores:
                    self.stores_cache[zip_code] = {}
                    for store in stores:
                        store_number = store.get('storeNumber')
                        division_number = store.get('divisionNumber')
                        self.stores_cache[zip_code][store_number] = division_number, store_number
                        if None not in self.stores_cache[zip_code]:
                            self.stores_cache[zip_code][None] = division_number, store_number
        return self.stores_cache.get(zip_code, {}).get(store_id, (None, None))

    def _get_products_info(self, upcs, division_id, store_id):
        products_info = []

        product_url = 'https://www.kroger.com/products/api/products/details'

        product_params = {'filterBadProducts': False, 'upcs': upcs}

        product_cookies = {'DivisionID': str(division_id), 'StoreCode': str(store_id)}

        response = requests.post(product_url, json=product_params, headers=self.headers,
                                 cookies=product_cookies, timeout=60)

        if self._check_response(response):
            data = response.json()

            products = data.get('products', [])

            for product in products:
                name = self._encode_unicode(product.get('description'))

                price = product.get('priceSale') or product.get('priceNormal')
                if price:
                    try:
                        price = float(price)
                    except ValueError:
                        price = 0
                else:
                    price = 0

                products_info.append({
                    'url': 'https://www.kroger.com/p/{}/{}'.format(product.get('slug'), product.get('upc')),
                    'name': name,
                    'availability': product.get('soldInStore'),
                    'size': product.get('customerFacingSize'),
                    'price': price,
                })

        return products_info

    def _get_shelf_info(self, shelf_id, division_id, store_id, brands=None):
        self.logger.debug('Checking store {}/{}: {}'.format(division_id, store_id, shelf_id))

        shelf_info = []

        shelf_url = 'https://www.kroger.com/search/api/searchAll'

        shelf_params = {
            'start': 0,
            'count': 100,
            'tab': 0,
            'taxonomyId': shelf_id
        }

        if brands:
            shelf_params['brandName'] = '|'.join(brands)

        shelf_cookies = {'DivisionID': str(division_id), 'StoreCode': str(store_id)}

        brands_checked = False

        while True:
            self.logger.info('Searching for products with params {}'.format(shelf_params))

            response = requests.post(shelf_url, params=shelf_params, headers=self.headers,
                                     cookies=shelf_cookies, timeout=60)

            if self._check_response(response):
                data = response.json()

                if brands and not brands_checked:
                    brands_lower = map(lambda x: x.lower(), brands)

                    params_changed = False
                    for shelf_brand in data.get('productsInfo', {}).get('brands', []):
                        shelf_brand = shelf_brand['name']
                        shelf_brand_lower = shelf_brand.lower()

                        try:
                            index = brands_lower.index(shelf_brand_lower)

                            if brands[index] != shelf_brand:
                                self.logger.info('Add correct brand filter: {}'.format(shelf_brand))
                                shelf_params['brandName'] += '|{}'.format(shelf_brand)
                                params_changed = True
                        except ValueError:
                            pass
                    brands_checked = True
                    if params_changed:
                        continue

                total = data.get('productsInfo', {}).get('totalCount', 0)

                upcs = data.get('upcs')

                if upcs:
                    shelf_info.extend(self._get_products_info(upcs, division_id, store_id))
                else:
                    dump_filename = uuid.uuid4().get_hex()

                    self.logger.warn(
                        'Empty items list, check dump: {}'.format(dump_filename))
                    self._save_response_dump(response, dump_filename)

                if shelf_params['start'] + shelf_params['count'] < total:
                    shelf_params['start'] += shelf_params['count']
                else:
                    break
            else:
                break

        return shelf_info

    def task_geo_competitor_assortment(self, options):
        self._check_options(options, params=['urls', 'stores'])

        stores = options.get('stores', [])

        failed_urls = []

        for url in options.get('urls', []):
            results = {}

            try:
                shelf_id = self._get_shelf_id(url)

                if shelf_id:
                    for store in stores:
                        zip_code, division_id, store_id = self._parse_store(store)

                        if not division_id or not store_id:
                            self.logger.warn('Cannot get store for zip code {} and store id {}'.format(zip_code, store_id))
                            continue

                        self.logger.info('Loading info for division id {} and store id {}'.format(division_id, store_id))

                        shelf_brands = self._get_shelf_brands(shelf_id, division_id, store_id)
                        if shelf_brands:
                            results[store_id] = shelf_brands
                        else:
                            self.logger.warn('Empty results for store: {}'.format(store_id))

                    all_brands = reduce(lambda x, y: x | set(y.keys()), results.values(), set())
                    all_brands = sorted(all_brands)

                    brands_filter = options.get('brands')
                    if brands_filter:
                        brands_filter = map(lambda x: x.lower(), brands_filter)

                        all_brands_lower = map(lambda x: x.lower(), all_brands)

                        filtered_brands = []

                        for filter_item in brands_filter:
                            try:
                                index = all_brands_lower.index(filter_item)
                                all_brands_lower.pop(index)

                                filtered_brands.append(all_brands.pop(index))
                            except ValueError:
                                pass

                        if '*' in brands_filter:
                            # add the rest brands
                            filtered_brands.extend(all_brands)

                        all_brands = filtered_brands

                    report_name = self._get_shelf_name(url)

                    style_red = PatternFill(fgColor=colors.RED, fill_type='solid')
                    style_green = PatternFill(fgColor=colors.GREEN, fill_type='solid')

                    if brands_filter and options.get('brands_comparison'):
                        report = Workbook(write_only=True)
                        sheet = report.create_sheet('Geo Competitor Comparison')

                        sheet.append(['Store ID', 'Primary Brand', 'Primary Brand # SKUs', 'Competitor Brand',
                                      'Competitor # SKUs'])

                        if all_brands:
                            primary_brand = all_brands[0]

                            for store_id, brands in results.iteritems():
                                primary_brand_count = brands.get(primary_brand, 0)

                                for brand in all_brands[1:]:
                                    brand_count = brands.get(brand, 0)

                                    primary_brand_cell = WriteOnlyCell(sheet, value=primary_brand_count)
                                    if primary_brand_count < brand_count:
                                        primary_brand_cell.fill = style_red
                                    elif primary_brand_count > brand_count:
                                        primary_brand_cell.fill = style_green

                                    sheet.append([store_id, primary_brand, primary_brand_cell, brand, brand_count])
                        else:
                            self.logger.warn('No brands after filter')

                        report.save(self.get_file_path_for_result('{}_comparison.xlsx'.format(report_name)))

                    report = Workbook(write_only=True)
                    sheet = report.create_sheet('Geo Competitor Assortment')

                    sheet.append(['Store ID'] + all_brands)

                    if brands_filter and all_brands:
                        primary_brand = all_brands[0]
                        all_brands = all_brands[1:]

                        for store_id, brands in results.iteritems():
                            primary_brand_count = results[store_id].get(primary_brand, 0)

                            primary_brand_cell = WriteOnlyCell(sheet, value=primary_brand_count)
                            if any(results[store_id].get(brand, 0) > primary_brand_count for brand in all_brands):
                                primary_brand_cell.fill = style_red
                            elif all(results[store_id].get(brand, 0) < primary_brand_count for brand in all_brands):
                                primary_brand_cell.fill = style_green

                            sheet.append([store_id, primary_brand_cell] +
                                         [brands.get(brand, 0) for brand in all_brands])
                    else:
                        for store_id, brands in results.iteritems():
                            sheet.append([store_id] + [brands.get(brand, 0) for brand in all_brands])

                    report.save(self.get_file_path_for_result('{}.xlsx'.format(report_name)))
                else:
                    raise Exception('Not shelf page: {}'.format(url))
            except:
                self.logger.error('Cannot process url: {}'.format(traceback.format_exc()))
                if url not in failed_urls:
                    failed_urls.append(url)

        if failed_urls:
            self.save_failed_urls(failed_urls)
            raise SitemapSpiderError('Some urls cannot be processed')

    def _get_shelf_id(self, url):
        shelf_parts = urlparse.urlparse(url)
        if shelf_parts.path.startswith('/pl/'):
            return shelf_parts.path.rsplit('/', 1)[-1]
        elif shelf_parts.path.startswith('/storecatalog/clicklistbeta'):
            parameters = urlparse.parse_qs(
                urlparse.urlparse(shelf_parts.fragment).query)
            return parameters.get('categoryId', [None])[0]

    def _get_shelf_name(self, url):
        shelf_parts = urlparse.urlparse(url)
        if shelf_parts.path.startswith('/pl/'):
            return shelf_parts.path.rsplit('/')[-2]
        elif shelf_parts.path.startswith('/storecatalog/clicklistbeta'):
            parameters = urlparse.parse_qs(
                urlparse.urlparse(shelf_parts.fragment).query)
            return parameters.get('pageTitle', [None])[0]

    def _get_shelf_brands(self, shelf_id, division_id, store_id):
        self.logger.debug('Checking store {}/{}: {}'.format(division_id, store_id, shelf_id))

        shelf_url = 'https://www.kroger.com/search/api/searchAll?start=0&count=24&tab=0&taxonomyId={shelf_id}'.format(
            shelf_id=shelf_id)

        shelf_cookies = {'DivisionID': str(division_id), 'StoreCode': str(store_id)}

        response = requests.post(shelf_url, headers=self.headers, cookies=shelf_cookies, timeout=60)

        if self._check_response(response):
            data = response.json()

            return {b['name']: b['count'] for b in data.get('productsInfo', {}).get('brands', [])}

    def task_geo_pricing_report(self, options):
        self._check_options(options, params=['urls', 'stores', 'brands'])

        stores = options.get('stores', [])
        brands = options.get('brands', [])
        sizes = options.get('sizes', []) or ['All Sizes']
        sizes_variants = {}

        for index, size in enumerate(sizes):
            if isinstance(size, list):
                sizes_variants[size[0]] = map(lambda x: x.lower(), size)
                sizes[index] = size[0]

        failed_urls = []

        for url in options.get('urls', []):
            results = {}

            try:
                shelf_id = self._get_shelf_id(url)

                if shelf_id:
                    for store in stores:
                        zip_code, division_id, store_id = self._parse_store(store)

                        if not division_id or not store_id:
                            self.logger.warn('Cannot get store for zip code {} and store id {}'.format(
                                zip_code, store_id))
                            continue

                        for brand in brands:
                            self.logger.info('Loading info for store id: {}/{}, brand: {}'.format(
                                division_id, store_id, self._encode_unicode(brand)))

                            shelf_info = self._get_shelf_info(shelf_id, division_id, store_id, [brand])
                            if shelf_info:
                                for size in sizes:
                                    if size == 'All Sizes':
                                        results.setdefault(store_id, {}).setdefault(brand, {})[size] = shelf_info
                                    elif sizes_variants.get(size):
                                        # check product name
                                        size_shelf_info = filter(lambda x: any(str(v) in (x['size'] or '').lower()
                                                                               for v in sizes_variants[size]),
                                                                 shelf_info)

                                        results.setdefault(store_id, {}).setdefault(brand, {})[size] = size_shelf_info
                                    else:
                                        size_filter = re.search('(\d+)', size)
                                        if size_filter:
                                            size_filter = size_filter.group(1)

                                        for product_info in shelf_info:
                                            product_size = re.search('(\d+)[\w\s]*/', product_info['size'])
                                            if not product_size or product_size.group(1) != size_filter:
                                                continue

                                            results.setdefault(store_id, {}).setdefault(brand, {}).setdefault(
                                                size, []).append(product_info)
                            else:
                                self.logger.warn('Empty results')

                    report = Workbook(write_only=True)
                    sheet = report.create_sheet('Geo Pricing Report')

                    sheet.append([None] + sizes)

                    for brand in brands:
                        size_prices = []

                        for size in sizes:
                            prices = []

                            for store_result in results.values():
                                prices.extend(x.get('price', 0) for x in store_result.get(brand, {}).get(size, []))

                            prices = filter(None, prices)
                            average_price = round(float(sum(prices)) / max(len(prices), 1), 2)

                            size_prices.append(average_price if average_price else None)

                        sheet.append([brand] + size_prices)

                    sheet = report.create_sheet('Raw Data')

                    style_green = PatternFill(fgColor=colors.GREEN, fill_type='solid')

                    header_1 = ['Store ID']
                    header_2 = [None]

                    for size in sizes:
                        header_1.extend([size] + [None] * (len(brands) - 1))
                        header_2.extend(brands)

                    sheet.append(header_1)
                    sheet.append(header_2)

                    for store_id, store_result in results.iteritems():
                        prices = []
                        num_of_rows = 0

                        for size in sizes:
                            size_prices = []

                            for brand in brands:
                                brand_prices = [x.get('price', None) for x in store_result.get(brand, {}).get(size, [])]

                                num_of_rows = max(num_of_rows, len(brand_prices))

                                size_prices.append(brand_prices)

                            if size_prices:
                                # color the lowest price
                                size_prices_all = filter(None, reduce(lambda x, y: x + y, size_prices, []))

                                if size_prices_all:
                                    size_prices_min = min(size_prices_all)

                                    primary_brand_prices = size_prices[0]

                                    for index, primary_brand_price in enumerate(primary_brand_prices):
                                        if primary_brand_price == size_prices_min:
                                            primary_brand_cell = WriteOnlyCell(sheet, value=primary_brand_price)
                                            primary_brand_cell.fill = style_green

                                            primary_brand_prices[index] = primary_brand_cell

                            prices.append(size_prices)

                        for row_index in range(num_of_rows):
                            row = [store_id if row_index == 0 else None]

                            for size_prices in prices:
                                for brand_prices in size_prices:
                                    row.append(brand_prices[row_index] if row_index < len(brand_prices) else None)

                            sheet.append(row)

                    sheet = report.create_sheet('Raw Data - Avg Prices')

                    sheet.append(header_1)
                    sheet.append(header_2)

                    for store_id, store_result in results.iteritems():
                        row = [store_id]

                        for size in sizes:
                            size_prices = []

                            for brand in brands:
                                prices = [x.get('price', None) for x in store_result.get(brand, {}).get(size, [])]

                                prices = filter(None, prices)
                                average_price = round(float(sum(prices)) / max(len(prices), 1), 2)

                                size_prices.append(average_price if average_price else None)

                            size_prices_all = filter(None, size_prices)

                            if size_prices_all and size_prices[0] == min(size_prices_all):
                                primary_brand_cell = WriteOnlyCell(sheet, value=size_prices[0])
                                primary_brand_cell.fill = style_green

                                size_prices[0] = primary_brand_cell

                            row.extend(size_prices)

                        sheet.append(row)

                    sheet = report.create_sheet('Raw Data (Rows)')

                    sheet.append(['Store ID', 'Product Name', 'Brand', 'Price'])

                    for store_id, store_result in results.iteritems():
                        for brand, brand_result in store_result.iteritems():
                            for size_result in brand_result.itervalues():
                                for product in size_result:
                                    sheet.append([store_id, product['name'], brand, product['price']])

                    report_name = self._get_shelf_name(url)
                    report.save(self.get_file_path_for_result('{}.xlsx'.format(report_name)))
                else:
                    raise Exception('Not shelf page: {}'.format(url))
            except:
                self.logger.error('Cannot process url: {}'.format(traceback.format_exc()))
                if url not in failed_urls:
                    failed_urls.append(url)

        if failed_urls:
            self.save_failed_urls(failed_urls)
            raise SitemapSpiderError('Some urls cannot be processed')

    def task_geo_out_of_stock_assortment(self, options):
        self._check_options(options, params=['urls', 'stores', 'brands'])

        stores = options.get('stores', [])
        brands = options.get('brands', [])

        failed_urls = []

        for url in options.get('urls', []):
            try:
                shelf_id = self._get_shelf_id(url)

                if shelf_id:
                    results = {}
                    all_results = {}
                    for store in stores:
                        zip_code, division_id, store_id = self._parse_store(store)

                        if not division_id or not store_id:
                            self.logger.warn('Cannot get store for zip code {} and store id {}'.format(
                                zip_code, store_id))
                            continue

                        if store_id in results:
                            continue

                        for brand in brands:
                            self.logger.info('Loading info for store id: {}/{}, brand: {}'.format(
                                division_id, store_id, self._encode_unicode(brand)))

                            shelf_info = self._get_shelf_info(shelf_id, division_id, store_id, [brand])

                            if shelf_info:
                                in_stock = 0
                                out_of_stock = 0
                                for product in shelf_info:
                                    if product['availability']:
                                        in_stock += 1
                                    else:
                                        out_of_stock += 1
                                total = in_stock + out_of_stock
                                results.setdefault(store_id, {})[brand] = [in_stock, out_of_stock, total]
                                all_results.setdefault(brand, {}).setdefault('in_stock', []).append(in_stock)
                                all_results[brand].setdefault('out_of_stock', []).append(out_of_stock)
                                all_results[brand].setdefault('total', []).append(total)
                            else:
                                self.logger.warn('Empty results')

                    report = Workbook(write_only=True)
                    sheet = report.create_sheet('Geo Out of Stock Assortment')

                    header = [None]
                    average = ['Average']

                    for brand in brands:
                        header.extend([brand, None, None, None])
                        if brand in all_results:
                            in_stock_average = float(sum(all_results[brand]['in_stock'])) / len(
                                all_results[brand]['in_stock'])
                            out_of_stock_average = float(sum(all_results[brand]['out_of_stock'])) / len(
                                all_results[brand]['out_of_stock'])
                            total_average = float(sum(all_results[brand]['total'])) / len(
                                all_results[brand]['total'])
                        else:
                            in_stock_average = 0
                            out_of_stock_average = 0
                            total_average = 0
                        average.extend([in_stock_average, out_of_stock_average, total_average, None])

                    sheet.append(header)
                    sheet.append(average)
                    sheet.append([])
                    sheet.append(['Store ID'] + ['In Stock', 'Out of Stock', 'Total', None] * len(brands))

                    for store_id, store_result in results.iteritems():
                        row = [store_id]

                        for brand in brands:
                            row.extend(store_result.get(brand, [0, 0, 0]) + [None])

                        sheet.append(row)

                    report_name = self._get_shelf_name(url)
                    report.save(self.get_file_path_for_result('{}.xlsx'.format(report_name)))
                else:
                    raise Exception('Not shelf page: {}'.format(url))
            except:
                self.logger.error('Cannot process url: {}'.format(traceback.format_exc()))
                if url not in failed_urls:
                    failed_urls.append(url)

        if failed_urls:
            self.save_failed_urls(failed_urls)
            raise SitemapSpiderError('Some urls cannot be processed')

    def task_shelf_to_item_urls(self, options):
        self._check_options(options, params=['urls'])

        shelf_urls = options.get('urls', [])

        session = requests.session()
        session.headers.update(self.headers)
        response = session.get('https://www.kroger.com/products/api/next-basket', timeout=60)
        self._check_response(response, raise_error=True, session=session)

        failed_urls = []
        for shelf_url in shelf_urls:
            try:
                item_urls_filename = '{}.csv'.format(
                    self._url_to_filename(shelf_url if isinstance(shelf_urls, list) else shelf_urls[shelf_url]))

                taxonomyId = self._get_shelf_id(shelf_url)
                if not taxonomyId:
                    raise SitemapSpiderError('Cannot find taxonomy Id in {}'.format(shelf_url))

                with open(self.get_file_path_for_result(item_urls_filename), 'w') as item_urls_file:
                    item_urls_writer = csv.writer(item_urls_file)

                    params = {
                        'start': 0,
                        'count': 100,
                        'tab': 0,
                        'taxonomyId': taxonomyId
                    }

                    while True:
                        self.logger.info('Searching for products with params {}'.format(params))

                        response = session.post('https://www.kroger.com/search/api/searchAll', params=params, timeout=60)

                        self._check_response(response, raise_error=True, session=session)

                        data = response.json()

                        total = data.get('productsInfo', {}).get('totalCount', 0)

                        upcs = data.get('upcs')

                        if upcs:
                            self.logger.info('Loading {} UPCs: {}'.format(len(upcs), upcs))

                            try:
                                response = session.post('https://www.kroger.com/products/api/products/details',
                                                        json={'filterBadProducts': False, 'upcs': upcs}, timeout=60)

                                self._check_response(response, raise_error=True, session=session)

                                products = response.json().get('products', [])

                                self.logger.info('Got {} products'.format(len(products)))

                                for product in products:
                                    slug = product.get('slug')
                                    upc = product.get('upc')
                                    item_url = 'https://www.kroger.com/p/{}/{}'.format(slug, upc)
                                    item_urls_writer.writerow([item_url])
                            except:
                                self.logger.error('Could not load products: {}'.format(traceback.format_exc()))
                        else:
                            dump_filename = uuid.uuid4().get_hex()

                            self.logger.warn('Empty items list, check dump: {}'.format(dump_filename))
                            self._save_response_dump(response, dump_filename)

                        if params['start'] + params['count'] < total:
                            params['start'] += params['count']
                        else:
                            break
            except:
                self.logger.error('Cannot process url: {}'.format(traceback.format_exc()))
                failed_urls.append(shelf_url)
        if failed_urls:
            self.save_failed_urls(failed_urls)
            raise SitemapSpiderError('Some urls cannot be processed')

    def task_stores(self, options):
        workers, tasks, output = self._start_workers(self._parse_stores, count=10)

        for store_url in self._parse_sitemap(self.STORES_SITEMAP_URL, headers=self.headers):
            tasks.put({'url': store_url})

        stores_filename = 'kroger_stores_{}.csv'.format(datetime.now().strftime('%Y-%m-%d'))

        with open(self.get_file_path_for_result(stores_filename), 'w') as stores_file:
            stores_csv = csv.writer(stores_file)
            stores_csv.writerow(['Zip Code', 'Store ID'])

            while True:
                try:
                    result = output.get(block=True, timeout=60)
                    zip_code = result.get('zip_code')
                    if zip_code:
                        stores_csv.writerow([zip_code, result.get('store_id')])
                except:
                    self.logger.info('Finish')
                    break

        self._stop_workers(workers, tasks)

    def _parse_stores(self, tasks, output):
        for task in iter(tasks.get, 'STOP'):
            store_url = task.get('url')

            store_id = re.search(r'/stores/details/\d+/(\d+)', store_url)

            if store_id:
                store_id = store_id.group(1)
                zip_code = None

                self.logger.info('Loading store: {}'.format(store_url))

                for i in range(3):
                    try:
                        response = requests.get(store_url, headers=self.headers, timeout=60)

                        self._check_response(response, raise_error=True)

                        if 'store location information cannot be found at this time' in response.content:
                            self.logger.warn('{}: store location information cannot be found at this time'.format(
                                store_id))
                            break

                        html = etree.HTML(response.content)

                        json_data = html.xpath('.//script[@type="application/ld+json"]/text()')
                        if json_data:
                            parser = HTMLParser()
                            store_details = json.loads(parser.unescape(json_data[0]))
                            zip_code = store_details.get('address', {}).get('postalCode')
                        if not zip_code:
                            self.logger.warn('No zip code for store {}'.format(store_id))
                            continue
                    except:
                        self.logger.error('Can not load store {}'.format(store_id))
                    else:
                        break

                result = {'store_id': store_id, 'zip_code': zip_code}

                output.put(result)
