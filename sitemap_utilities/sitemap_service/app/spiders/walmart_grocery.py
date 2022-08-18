import csv
import re
import traceback
import time
import requests

from urlparse import urlparse

from openpyxl import Workbook
from openpyxl.worksheet.write_only import WriteOnlyCell
from openpyxl.styles import PatternFill, colors

from . import SitemapSpider, SitemapSpiderError


class WalmartGrocerySitemapSpider(SitemapSpider):
    retailer = 'grocery.walmart.com'

    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.12; rv:54.0) Gecko/20100101 Firefox/54.0'
    }

    def __init__(self, *args, **kwargs):
        super(WalmartGrocerySitemapSpider, self).__init__(*args, **kwargs)

        self.stores_cache = {}
        self.departments_cache = {}
        self.brands_cache = {}

    def task_geo_report(self, options):
        required_options = ['stores']

        if not options.get('search_terms'):
            required_options.append('urls')

        self._check_options(options, params=required_options)

        report_name = options.get('request_name') or 'geo_report'

        with open(self.get_file_path_for_result('{}.csv'.format(report_name)), 'wb') as geo_report_file:
            csv_writer = csv.writer(geo_report_file)
            csv_writer.writerow(['Zip Code', 'Store ID', 'Product Name', 'Brand', 'URL', 'Availability', 'Price'])

            stores = options.get('stores', [])

            failed_urls = []
            for store in stores:
                store_id, zip_code = self._parse_store(store)

                if not store_id:
                    self.logger.warn('Missing store id for zip_code: {}'.format(zip_code))
                    continue

                self.logger.info('Loading info for store id: {}'.format(store_id))

                for url in options.get('urls', []):
                    try:
                        shelf_id = self._get_shelf_id(url)

                        if shelf_id:
                            shelf_info = self._get_shelf_info(shelf_id, store_id, options.get('brands'))

                            csv_writer.writerows([zip_code, store_id,
                                                  product_info.get('name'),
                                                  product_info.get('brand'),
                                                  product_info.get('url'),
                                                  product_info.get('availability'),
                                                  product_info.get('price')]
                                                 for product_info in shelf_info)
                        else:
                            product_info = self._get_product_info(url, store_id)

                            csv_writer.writerow([zip_code, store_id,
                                                 product_info.get('name'),
                                                 product_info.get('brand'),
                                                 url,
                                                 product_info.get('availability'),
                                                 product_info.get('price')])
                    except:
                        self.logger.error('Cannot process url: {}'.format(traceback.format_exc()))
                        if url not in failed_urls:
                            failed_urls.append(url)

                for search_term in options.get('search_terms', []):
                    search_info = self._get_search_info(search_term, store_id, options.get('department'),
                                                        options.get('brands'))

                    csv_writer.writerows([zip_code, store_id,
                                          product_info.get('name'),
                                          product_info.get('brand'),
                                          product_info.get('url'),
                                          product_info.get('availability'),
                                          product_info.get('price')]
                                         for product_info in search_info)

            if failed_urls:
                self.save_failed_urls(failed_urls)
                raise SitemapSpiderError('Some urls cannot be processed')

    def _get_store_id(self, zip_code):
        if zip_code not in self.stores_cache:
            self.logger.info('Search store for zip code: {}'.format(zip_code))

            store_search_url = 'https://grocery.walmart.com/v3/api/serviceAvailability?' \
                               'postalCode={zip_code}'.format(zip_code=zip_code)

            response = requests.get(store_search_url, headers=self.headers, timeout=60)

            if self._check_response(response):
                data = response.json()

                self.stores_cache[zip_code] = data.get('storeId')

        return self.stores_cache.get(zip_code)

    def _get_product_info(self, url, store_id):
        self.logger.debug('Checking store {}: {}'.format(store_id, url))

        product_info = {
            'name': None,
            'brand': None,
            'price': None,
            'availability': None
        }

        product_id = self._get_product_id(url)

        product_url = 'https://grocery.walmart.com/v3/api/products/{product_id}?' \
                      'itemFields=all&storeId={store_id}'.format(product_id=product_id, store_id=store_id)
        for i in range(3):
            try:
                response = requests.get(product_url, headers=self.headers, timeout=60)

                if self._check_response(response):
                    data = response.json()

                    name = self._encode_unicode(data.get('basic', {}).get('name'))
                    product_info['name'] = name

                    brand = self._encode_unicode(data.get('detailed', {}).get('brand'))
                    product_info['brand'] = brand

                    product_info['availability'] = data.get('store', {}).get('isInStock')
                    product_info['price'] = data.get('store', {}).get('price', {}).get('list')

                    if product_info['availability'] is None and product_info['price'] is None:
                        self.logger.info('No data. Try again in {} seconds'.format(i))
                        time.sleep(i)
                        continue
            except:
                self.logger.error('Product info error: {}'.format(traceback.format_exc()))
            else:
                break
        else:
            product_info['availability'] = 'Item No Longer Available'

        return product_info

    def _get_shelf_info(self, shelf_id, store_id, brands=None, sizes=None):
        self.logger.debug('Checking store {}: {}'.format(store_id, shelf_id))

        shelf_info = []

        page = 1
        offset = 0

        while page:
            self.logger.info('Page: {}'.format(page))

            shelf_url = 'https://grocery.walmart.com/v3/api/products?strategy=aisle&taxonomyNodeId={shelf_id}&' \
                        'storeId={store_id}&count=60&page={page}&offset={offset}'.format(shelf_id=shelf_id,
                                                                                         store_id=store_id,
                                                                                         page=page,
                                                                                         offset=offset)

            if brands:
                shelf_url += ''.join('&filter=ghs_brand~{}'.format(self._encode_unicode(brand)) for brand in brands)

            if sizes:
                shelf_url += ''.join('&filter=ghs_size~{}'.format(size) for size in sizes)

            for i in range(3):
                try:
                    response = requests.get(shelf_url, headers=self.headers, timeout=60)

                    if self._check_response(response):
                        data = response.json()

                        products = data.get('products')

                        if not products:
                            self.logger.info('No data. Try again in {} seconds'.format(i))

                            if i == 0:
                                filters = data.get('filters', [])

                                if brands:
                                    shelf_url = self._fix_brands_in_url(filters, brands, shelf_url)

                                if sizes:
                                    shelf_url = self._fix_sizes_in_url(filters, sizes, shelf_url)

                            time.sleep(i)
                            continue

                        shelf_info.extend(self._parse_products_info(products, store_id))

                        offset += 60

                        if offset < data.get('totalCount', 0):
                            page += 1
                        else:
                            page = None
                    else:
                        page = None
                except:
                    self.logger.error('Shelf info error: {}'.format(traceback.format_exc()))
                else:
                    break
            else:
                page = None

        return shelf_info

    def _get_search_info(self, search_term, store_id, department=None, brands=None, sizes=None):
        self.logger.debug('Searching store {}: {}'.format(store_id, search_term))

        search_info = []

        page = 1
        offset = 0

        while page:
            self.logger.info('Page: {}'.format(page))

            search_url = 'https://grocery.walmart.com/v3/api/products?strategy=search&query={search_term}&' \
                         'storeId={store_id}&count=60&page={page}&offset={offset}'.format(search_term=search_term,
                                                                                          store_id=store_id,
                                                                                          page=page,
                                                                                          offset=offset)

            if department:
                department_id = self._get_department_id(search_term, store_id, department)
                if department_id:
                    search_url += '&filter=Departments~{}'.format(department_id)
                else:
                    self.logger.warn('No match for department: {}'.format(department))

            if brands:
                search_url += ''.join('&filter=ghs_brand~{}'.format(brand) for brand in brands)

            if sizes:
                search_url += ''.join('&filter=ghs_size~{}'.format(size) for size in sizes)

            for i in range(3):
                try:
                    response = requests.get(search_url, headers=self.headers, timeout=60)

                    if self._check_response(response):
                        data = response.json()

                        products = data.get('products')

                        if not products:
                            self.logger.info('No data. Try again in {} seconds'.format(i))

                            if i == 0:
                                filters = data.get('searchFacets', {}).get('facets', [])

                                if brands:
                                    search_url = self._fix_brands_in_url(filters, brands, search_url)

                                if sizes:
                                    search_url = self._fix_sizes_in_url(filters, sizes, search_url)

                            time.sleep(i)
                            continue

                        search_info.extend(self._parse_products_info(products, store_id))

                        offset += 60

                        if offset < data.get('totalCount', 0):
                            page += 1
                        else:
                            page = None
                    else:
                        page = None
                except:
                    self.logger.error('Search info error: {}'.format(traceback.format_exc()))
                else:
                    break
            else:
                page = None

        return search_info

    def _get_department_id(self, search_term, store_id, department):
        key = search_term, store_id, department
        if key not in self.departments_cache:
            self.logger.debug('Searching department {} in {}'.format(department, store_id))

            search_url = 'https://grocery.walmart.com/v3/api/products?strategy=search&query={search_term}&' \
                         'storeId={store_id}&count=60&page=1&offset=0'.format(search_term=search_term,
                                                                              store_id=store_id)

            for i in range(3):
                try:
                    response = requests.get(search_url, headers=self.headers, timeout=60)

                    if self._check_response(response):
                        data = response.json()

                        departments = data.get('searchFacets', {}).get('departments', [])

                        if not departments:
                            self.logger.info('No data. Try again in {} seconds'.format(i))
                            time.sleep(i)
                            continue

                        department_id = self._search_department(department, departments)
                        if department_id:
                            self.departments_cache[key] = department_id
                except:
                    self.logger.error('Search department error: {}'.format(traceback.format_exc()))
                else:
                    break

        return self.departments_cache.get(key)

    def _search_department(self, department_name, departments):
        department_name = department_name.lower()
        for department in departments:
            if department.get('name', '').lower() == department_name:
                return department.get('id')
            children = department.get('children', [])
            if children:
                department_id = self._search_department(department_name, children)
                if department_id:
                    return department_id

    def _fix_brands_in_url(self, filters, brands, url):
        for f in filters:
            if f.get('name') == 'ghs_brand':
                brands_lower = map(lambda x: x.lower(), brands)

                for shelf_brand in f.get('values', []):
                    shelf_brand = shelf_brand['name']
                    shelf_brand_lower = shelf_brand.lower()

                    try:
                        index = brands_lower.index(shelf_brand_lower)

                        if brands[index] != shelf_brand:
                            shelf_brand = self._encode_unicode(shelf_brand)

                            self.logger.info('Add correct brand filter: {}'.format(shelf_brand))
                            url += '&filter=ghs_brand~{}'.format(shelf_brand)
                    except ValueError:
                        pass

                break
        return url

    def _fix_sizes_in_url(self, filters, sizes, url):
        for f in filters:
            if f.get('name') == 'ghs_size':
                sizes_lower = map(lambda x: x.lower(), sizes)

                for shelf_size in f.get('values', []):
                    shelf_size = shelf_size['name']
                    shelf_size_lower = shelf_size.lower()

                    try:
                        index = sizes_lower.index(shelf_size_lower)

                        if sizes[index] != shelf_size:
                            self.logger.info('Add correct size filter: {}'.format(shelf_size))
                            url += '&filter=ghs_size~{}'.format(shelf_size)
                    except ValueError:
                        pass

                break
        return url

    def _parse_products_info(self, products, store_id):
        products_info = []
        for product in products:
            name = self._encode_unicode(product.get('basic', {}).get('name'))

            product_id = product.get('USItemId')
            product_url = 'https://grocery.walmart.com/product/{}'.format(product_id)

            if product_id not in self.brands_cache:
                brand = self._get_product_info(product_url, store_id)['brand']
                self.brands_cache[product_id] = brand
            else:
                brand = self.brands_cache[product_id]

            product_info = {
                'url': product_url,
                'name': name,
                'brand': brand,
                'price': product.get('store', {}).get('price', {}).get('list'),
                'availability': product.get('store', {}).get('isInStock')
            }

            products_info.append(product_info)
        return products_info

    def _get_product_id(self, url):
        product_id = re.search(r'/product/(\d+)', url)

        if not product_id:
            product_id = re.search(r'skuId=(\d+)', url)

        if product_id:
            return product_id.group(1)

        return urlparse(url).path.split('/')[-1]

    def _get_shelf_id(self, url):
        shelf_id = re.search(r'aisle=([^&]*)', url)
        if shelf_id:
            return shelf_id.group(1)

    def _get_shelf_name(self, url):
        return urlparse(url).path.split('/')[-1]

    def task_geo_competitor_assortment(self, options):
        required_options = ['stores']

        if not options.get('search_terms'):
            required_options.append('urls')

        self._check_options(options, params=required_options)

        stores = options.get('stores', [])
        sizes = options.get('sizes', [])

        failed_urls = []

        for url in options.get('urls', []):
            results = {}

            try:
                shelf_id = self._get_shelf_id(url)

                if shelf_id:
                    for store in stores:
                        store_id, zip_code = self._parse_store(store)

                        if not store_id:
                            self.logger.warn('Missing store id for zip_code: {}'.format(zip_code))
                            continue

                        self.logger.info('Loading info for store id: {}'.format(store_id))

                        shelf_brands = self._get_shelf_brands(shelf_id, store_id, sizes)
                        if shelf_brands:
                            results[store_id] = shelf_brands
                        else:
                            self.logger.warn('Empty results for store: {}'.format(store_id))

                    report_name = self._get_shelf_name(url)

                    self._save_geo_competitor_assortment_report(report_name,
                                                                results,
                                                                options)
                else:
                    raise Exception('Not shelf page: {}'.format(url))
            except:
                self.logger.error('Cannot process url: {}'.format(traceback.format_exc()))
                if url not in failed_urls:
                    failed_urls.append(url)

        for search_term in options.get('search_terms', []):
            results = {}

            for store in stores:
                store_id, zip_code = self._parse_store(store)

                if not store_id:
                    self.logger.warn('Missing store id for zip_code: {}'.format(zip_code))
                    continue

                self.logger.info('Loading info for store id: {}'.format(store_id))

                search_brands = self._get_search_brands(search_term, store_id, options.get('department'), sizes)
                if search_brands:
                    results[store_id] = search_brands
                else:
                    self.logger.warn('Empty results for store: {}'.format(store_id))

            report_name = search_term.replace(' ', '_')

            self._save_geo_competitor_assortment_report(report_name, results, options)

        if failed_urls:
            self.save_failed_urls(failed_urls)
            raise SitemapSpiderError('Some urls cannot be processed')

    def _save_geo_competitor_assortment_report(self, report_name, results, options):
        all_brands = reduce(lambda x, y: x | set(y.keys()), results.values(), set())
        all_brands = sorted(all_brands)

        brands_filter = options.get('brands')
        if brands_filter:
            brands_filter = map(lambda x: x.lower(), brands_filter)

            all_brands_lower = map(lambda x: x.lower(), all_brands)

            filtered_brands = []

            for filter_item in brands_filter:
                try:
                    index = all_brands_lower.index(filter_item)
                    all_brands_lower.pop(index)

                    filtered_brands.append(all_brands.pop(index))
                except ValueError:
                    pass

            if '*' in brands_filter:
                # add the rest brands
                filtered_brands.extend(all_brands)

            all_brands = filtered_brands

        style_red = PatternFill(fgColor=colors.RED, fill_type="solid")
        style_green = PatternFill(fgColor=colors.GREEN, fill_type="solid")

        if brands_filter and options.get('brands_comparison'):
            report = Workbook(write_only=True)
            sheet = report.create_sheet('Geo Competitor Comparison')

            sheet.append(['Store ID', 'Primary Brand', 'Primary Brand # SKUs', 'Competitor Brand',
                          'Competitor # SKUs'])

            if all_brands:
                primary_brand = all_brands[0]

                for store_id, brands in results.iteritems():
                    primary_brand_count = brands.get(primary_brand, 0)

                    for brand in all_brands[1:]:
                        brand_count = brands.get(brand, 0)

                        primary_brand_cell = WriteOnlyCell(sheet, value=primary_brand_count)
                        if primary_brand_count < brand_count:
                            primary_brand_cell.fill = style_red
                        elif primary_brand_count > brand_count:
                            primary_brand_cell.fill = style_green

                        sheet.append([store_id, primary_brand, primary_brand_cell, brand, brand_count])
            else:
                self.logger.warn('No brands after filter')

            report.save(self.get_file_path_for_result('{}_comparison.xlsx'.format(report_name)))

        report = Workbook(write_only=True)
        sheet = report.create_sheet('Geo Competitor Assortment')

        sheet.append(['Store ID'] + all_brands)

        if brands_filter and all_brands:
            primary_brand = all_brands[0]
            all_brands = all_brands[1:]

            for store_id, brands in results.iteritems():
                primary_brand_count = results[store_id].get(primary_brand, 0)

                primary_brand_cell = WriteOnlyCell(sheet, value=primary_brand_count)
                if any(results[store_id].get(brand, 0) > primary_brand_count for brand in all_brands):
                    primary_brand_cell.fill = style_red
                elif all(results[store_id].get(brand, 0) < primary_brand_count for brand in all_brands):
                    primary_brand_cell.fill = style_green

                sheet.append([store_id, primary_brand_cell] +
                             [brands.get(brand, 0) for brand in all_brands])
        else:
            for store_id, brands in results.iteritems():
                sheet.append([store_id] + [brands.get(brand, 0) for brand in all_brands])

        report.save(self.get_file_path_for_result('{}.xlsx'.format(report_name)))

    def _get_shelf_brands(self, shelf_id, store_id, sizes=None):
        self.logger.debug('Checking store {}: {}'.format(store_id, shelf_id))

        shelf_url = 'https://grocery.walmart.com/v3/api/products?strategy=aisle&taxonomyNodeId={shelf_id}&' \
                    'storeId={store_id}&count=60&page=1&offset=0'.format(shelf_id=shelf_id, store_id=store_id)

        if sizes:
            shelf_url += ''.join('&filter=ghs_size~{}'.format(size) for size in sizes)

        for i in range(3):
            try:
                response = requests.get(shelf_url, headers=self.headers, timeout=60)

                if self._check_response(response):
                    data = response.json()

                    filters = data.get('filters', [])

                    if not filters:
                        self.logger.info('No data. Try again in {} seconds'.format(i))

                        if i == 0:
                            filters = data.get('filters', [])

                            if sizes:
                                shelf_url = self._fix_sizes_in_url(filters, sizes, shelf_url)

                        time.sleep(i)
                        continue

                    for f in filters:
                        if f.get('name') == 'ghs_brand':
                            return dict((x['name'], x['count']) for x in f.get('values', []))
            except:
                self.logger.error('Shelf brands error: {}'.format(traceback.format_exc()))

    def _get_search_brands(self, search_term, store_id, department=None, sizes=None):
        self.logger.debug('Searching store {}: {}'.format(store_id, search_term))

        search_url = 'https://grocery.walmart.com/v3/api/products?strategy=search&query={search_term}&' \
                     'storeId={store_id}&count=60&page=1&offset=0'.format(search_term=search_term,
                                                                          store_id=store_id)

        if department:
            department_id = self._get_department_id(search_term, store_id, department)
            if department_id:
                search_url += '&filter=Departments~{}'.format(department_id)
            else:
                self.logger.warn('No match for department: {}'.format(department))

        if sizes:
            search_url += ''.join('&filter=ghs_size~{}'.format(size) for size in sizes)

        for i in range(3):
            try:
                response = requests.get(search_url, headers=self.headers, timeout=60)

                if self._check_response(response):
                    data = response.json()

                    filters = data.get('searchFacets', {}).get('facets', [])

                    if not filters:
                        self.logger.info('No data. Try again in {} seconds'.format(i))

                        if i == 0:
                            filters = data.get('filters', [])

                            if sizes:
                                search_url = self._fix_sizes_in_url(filters, sizes, search_url)

                        time.sleep(i)
                        continue

                    for f in filters:
                        if f.get('name') == 'ghs_brand':
                            return dict((x['name'], x['count']) for x in f.get('values', []))
            except:
                self.logger.error('Shelf brands error: {}'.format(traceback.format_exc()))

    def _parse_store(self, store):
        zip_code = None
        store_id = None

        if isinstance(store, dict):
            zip_code = store.get('zip_code')
            store_id = store.get('store_id')
        elif isinstance(store, (list, tuple)):
            if len(store) == 1:
                store_id = store[0],
            elif len(store) > 1:
                zip_code = store[0],
                store_id = store[1]
        else:
            store_id = store

        if zip_code and not store_id:
            store_id = self._get_store_id(zip_code)

        return store_id, zip_code

    def task_geo_pricing_report(self, options):
        required_options = ['stores', 'brands']

        if not options.get('search_terms'):
            required_options.append('urls')

        self._check_options(options, params=required_options)

        stores = options.get('stores', [])
        brands = options.get('brands', [])
        sizes = options.get('sizes', []) or ['All Sizes']
        sizes_variants = {}

        for index, size in enumerate(sizes):
            if isinstance(size, list):
                sizes_variants[size[0]] = map(lambda x: x.lower(), size)
                sizes[index] = size[0]

        failed_urls = []

        for url in options.get('urls', []):
            results = {}

            try:
                shelf_id = self._get_shelf_id(url)

                if shelf_id:
                    for store in stores:
                        store_id, zip_code = self._parse_store(store)

                        if not store_id:
                            self.logger.warn('Missing store id for zip_code: {}'.format(zip_code))
                            continue

                        for brand in brands:
                            for size in sizes:
                                self.logger.info('Loading info for store id: {}, brand: {}, size: {}'.format(
                                    store_id, self._encode_unicode(brand), size))

                                shelf_info = self._get_shelf_info(
                                    shelf_id, store_id, [brand],
                                    [size] if not sizes_variants.get(size) and size != 'All Sizes' else None
                                )

                                if shelf_info:
                                    if sizes_variants.get(size):
                                        # check product name
                                        shelf_info = filter(lambda x: any(str(v) in (x['name'] or '').lower()
                                                                          for v in sizes_variants[size]),
                                                            shelf_info)

                                    results.setdefault(store_id, {}).setdefault(brand, {})[size] = shelf_info
                                else:
                                    self.logger.warn('Empty results')

                    report_name = self._get_shelf_name(url)

                    self._save_geo_pricing_report(report_name, results, brands, sizes)
                else:
                    raise Exception('Not shelf page: {}'.format(url))
            except:
                self.logger.error('Cannot process url: {}'.format(traceback.format_exc()))
                if url not in failed_urls:
                    failed_urls.append(url)

        for search_term in options.get('search_terms', []):
            results = {}

            for store in stores:
                store_id, zip_code = self._parse_store(store)

                if not store_id:
                    self.logger.warn('Missing store id for zip_code: {}'.format(zip_code))
                    continue

                for brand in brands:
                    for size in sizes:
                        self.logger.info('Loading search for store id: {}, brand: {}, size: {}'.format(
                            store_id, brand, size))

                        search_info = self._get_search_info(
                            search_term, store_id, options.get('department'), [brand],
                            [size] if not sizes_variants.get(size) and size != 'All Sizes' else None
                        )

                        if search_info:
                            if sizes_variants.get(size):
                                # check product name
                                search_info = filter(lambda x: any(str(v) in (x['name'] or '').lower()
                                                                   for v in sizes_variants[size]),
                                                     search_info)

                            results.setdefault(store_id, {}).setdefault(brand, {})[size] = search_info
                        else:
                            self.logger.warn('Empty results')

            report_name = search_term.replace(' ', '_')

            self._save_geo_pricing_report(report_name, results, brands, sizes)

        if failed_urls:
            self.save_failed_urls(failed_urls)
            raise SitemapSpiderError('Some urls cannot be processed')

    def _save_geo_pricing_report(self, report_name, results, brands, sizes):
        report = Workbook(write_only=True)
        sheet = report.create_sheet('Geo Pricing Report')

        sheet.append([None] + sizes)

        for brand in brands:
            size_prices = []

            for size in sizes:
                prices = []

                for store_result in results.values():
                    prices.extend(x.get('price', 0) for x in store_result.get(brand, {}).get(size, []))

                prices = filter(None, prices)
                average_price = round(float(sum(prices)) / max(len(prices), 1), 2)

                size_prices.append(average_price if average_price else None)

            sheet.append([brand] + size_prices)

        sheet = report.create_sheet('Raw Data')

        style_green = PatternFill(fgColor=colors.GREEN, fill_type="solid")

        header_1 = ['Store ID']
        header_2 = [None]

        for size in sizes:
            header_1.extend([size] + [None] * (len(brands) - 1))
            header_2.extend(brands)

        sheet.append(header_1)
        sheet.append(header_2)

        for store_id, store_result in results.iteritems():
            prices = []
            num_of_rows = 0

            for size in sizes:
                size_prices = []

                for brand in brands:
                    brand_prices = [x.get('price', None) for x in store_result.get(brand, {}).get(size, [])]

                    num_of_rows = max(num_of_rows, len(brand_prices))

                    size_prices.append(brand_prices)

                if size_prices:
                    # color the lowest price
                    size_prices_all = filter(None, reduce(lambda x, y: x + y, size_prices, []))

                    if size_prices_all:
                        size_prices_min = min(size_prices_all)

                        primary_brand_prices = size_prices[0]

                        for index, primary_brand_price in enumerate(primary_brand_prices):
                            if primary_brand_price == size_prices_min:
                                primary_brand_cell = WriteOnlyCell(sheet, value=primary_brand_price)
                                primary_brand_cell.fill = style_green

                                primary_brand_prices[index] = primary_brand_cell

                prices.append(size_prices)

            for row_index in range(num_of_rows):
                row = [store_id if row_index == 0 else None]

                for size_prices in prices:
                    for brand_prices in size_prices:
                        row.append(brand_prices[row_index] if row_index < len(brand_prices) else None)

                sheet.append(row)

        sheet = report.create_sheet('Raw Data - Avg Prices')

        sheet.append(header_1)
        sheet.append(header_2)

        for store_id, store_result in results.iteritems():
            row = [store_id]

            for size in sizes:
                size_prices = []

                for brand in brands:
                    prices = [x.get('price', None) for x in store_result.get(brand, {}).get(size, [])]

                    prices = filter(None, prices)
                    average_price = round(float(sum(prices)) / max(len(prices), 1), 2)

                    size_prices.append(average_price if average_price else None)

                size_prices_all = filter(None, size_prices)

                if size_prices_all and size_prices[0] == min(size_prices_all):
                    primary_brand_cell = WriteOnlyCell(sheet, value=size_prices[0])
                    primary_brand_cell.fill = style_green

                    size_prices[0] = primary_brand_cell

                row.extend(size_prices)

            sheet.append(row)

        sheet = report.create_sheet('Raw Data (Rows)')

        sheet.append(['Store ID', 'Product Name', 'Brand', 'Price'])

        for store_id, store_result in results.iteritems():
            for brand, brand_result in store_result.iteritems():
                for size_result in brand_result.itervalues():
                    for product in size_result:
                        sheet.append([store_id, product['name'], brand, product['price']])

        report.save(self.get_file_path_for_result('{}.xlsx'.format(report_name)))

    def task_geo_out_of_stock_assortment(self, options):
        self._check_options(options, params=['urls', 'stores', 'brands'])

        stores = options.get('stores', [])
        brands = options.get('brands', [])

        failed_urls = []

        for url in options.get('urls', []):
            try:
                shelf_id = self._get_shelf_id(url)

                if shelf_id:
                    results = {}
                    all_results = {}
                    for store in stores:
                        store_id, zip_code = self._parse_store(store)

                        if not store_id:
                            self.logger.warn('Missing store id for zip_code: {}'.format(zip_code))
                            continue

                        if store_id in results:
                            continue

                        for brand in brands:
                            self.logger.info('Loading info for store id: {}, brand: {}'.format(
                                store_id, self._encode_unicode(brand)))

                            shelf_info = self._get_shelf_info(shelf_id, store_id, [brand])

                            if shelf_info:
                                in_stock = 0
                                out_of_stock = 0
                                for product in shelf_info:
                                    if product['availability']:
                                        in_stock += 1
                                    else:
                                        out_of_stock += 1
                                total = in_stock + out_of_stock
                                results.setdefault(store_id, {})[brand] = [in_stock, out_of_stock, total]
                                all_results.setdefault(brand, {}).setdefault('in_stock', []).append(in_stock)
                                all_results[brand].setdefault('out_of_stock', []).append(out_of_stock)
                                all_results[brand].setdefault('total', []).append(total)
                            else:
                                self.logger.warn('Empty results')

                    report = Workbook(write_only=True)
                    sheet = report.create_sheet('Geo Out of Stock Assortment')

                    header = [None]
                    average = ['Average']

                    for brand in brands:
                        header.extend([brand, None, None, None])
                        if brand in all_results:
                            in_stock_average = float(sum(all_results[brand]['in_stock'])) / len(
                                all_results[brand]['in_stock'])
                            out_of_stock_average = float(sum(all_results[brand]['out_of_stock'])) / len(
                                all_results[brand]['out_of_stock'])
                            total_average = float(sum(all_results[brand]['total'])) / len(
                                all_results[brand]['total'])
                        else:
                            in_stock_average = 0
                            out_of_stock_average = 0
                            total_average = 0
                        average.extend([in_stock_average, out_of_stock_average, total_average, None])

                    sheet.append(header)
                    sheet.append(average)
                    sheet.append([])
                    sheet.append(['Store ID'] + ['In Stock', 'Out of Stock', 'Total', None] * len(brands))

                    for store_id, store_result in results.iteritems():
                        row = [store_id]

                        for brand in brands:
                            row.extend(store_result.get(brand, [0, 0, 0]) + [None])

                        sheet.append(row)

                    report_name = self._get_shelf_name(url)
                    report.save(self.get_file_path_for_result('{}.xlsx'.format(report_name)))
                else:
                    raise Exception('Not shelf page: {}'.format(url))
            except:
                self.logger.error('Cannot process url: {}'.format(traceback.format_exc()))
                if url not in failed_urls:
                    failed_urls.append(url)

        if failed_urls:
            self.save_failed_urls(failed_urls)
            raise SitemapSpiderError('Some urls cannot be processed')
