import xlsxwriter
import KohlsScraperAux
import downloadS3
import re
import logging

from urlparse import urlparse
from openpyxl import load_workbook


logger = logging.getLogger('mediaaudit')


def get_column_index(regex,  columns):
    for index, column in enumerate(columns):
        if column and re.search(regex, column):
            return index


def run(input_file, output_file):
    # All data goes into here #
    input_data = {}
    seasons = ['F17', 'S17', 'F16', 'S16', 'F15', 'F14']
    logger.debug("Kohls audit started.")
    # May need to re-name Sheet1 field #
    wb = load_workbook(filename=input_file, read_only=True)
    ws = wb['Sheet1']

    # Reading header
    columns = [cell.value.lower() if isinstance(cell.value, basestring) else cell.value for cell in ws[1]]

    web_id_index = get_column_index(r'web', columns)
    if web_id_index is None:
        web_id_index = 0

    pc9_index = get_column_index(r'pc9', columns)
    if pc9_index is None:
        pc9_index = 1

    color_index = get_column_index(r'wash|color', columns)
    if color_index is None:
        color_index = 2

    logger.debug("Input xlsx loaded.")
    # Reading in from File Mapping
    for row in ws.iter_rows(min_row=2):
        pc9 = row[pc9_index].value
        if pc9:
            input_data[pc9] = {
                'web_id': row[web_id_index].value,
                'color': row[color_index].value
            }
    logger.debug("Input xlsx loaded. Rows count = {0}".format(len(input_data.keys())))

    # Reading in from File Mapping
    for row in ws.iter_rows(min_row=2):
        pc9 = row[pc9_index].value
        if pc9:
            input_data[pc9] = {
                'web_id': row[web_id_index].value,
                'color': row[color_index].value.replace(' ', "_").lower()
                if isinstance(row[color_index].value, basestring) else row[color_index].value
            }

    #  We use a set to avoid duplicate Product IDs #
    products_cache = {}

    logger.debug("Products crawling started.")
    for pc9, data in input_data.iteritems():
        web_id = data['web_id']

        url = 'http://www.kohls.com/product/prd-{}/null.jsp'.format(web_id)
        if web_id not in products_cache:
            products_cache[web_id] = KohlsScraperAux.load_product(url)

        product = products_cache[web_id]
        product_state_check = KohlsScraperAux.check_state(product)
        logger.debug("Product {0} loaded. Loading state {1}".format(url, product.get('status')))
        if product and product_state_check:
            data['color_urls'] = KohlsScraperAux.get_color_urls(product)
            data['image_urls'] = KohlsScraperAux.get_image_urls(product)
            data['video'] = KohlsScraperAux.get_video(product)
            data['size_chart'] = KohlsScraperAux.get_size_chart(product)
        data['product_loaded'] = product_state_check

    logger.debug("Front images comparison started.")
    for pc9, data in input_data.iteritems():
        product_loaded = data['product_loaded']
        if not product_loaded:
            data['color_matched'] = 'Item not found'
            continue
        color_urls = data.get('color_urls')
        logger.debug('Processing product web_id={0}'.format(data.get('web_id')))

        data['color_found'] = True  # if front image not found we should skip alt images
        if color_urls:
            for color_url in color_urls:
                color_url_parts = urlparse(color_url)
                if "ALT" not in color_url_parts.path:
                    #  Make sure we are getting the right color #
                    color = color_url_parts.path.split("_", 1)[-1].lower()
                    if color == data['color']:
                        try:
                            logger.debug("Image match by color name found ({0}). Trying to compare..".format(color_url))
                            match_result = downloadS3.getComparison(pc9, seasons, 'F', color_url)
                        except Exception as e:
                            logger.error("Image matching error: {0}".format(e.message))
                            match_result = 0
                        logger.debug("Matching result = {0}".format(match_result))
                        if match_result > 90:
                            data['color_matched'] = 'Yes'
                        elif match_result > 70:
                            data['color_matched'] = 'Yes (cropped)'
                        elif match_result > 50:
                            data['color_matched'] = 'Indeterminate'
                        elif match_result == -1:
                            data['color_matched'] = 'S3 image not existed for that product'
                        else:
                            data['color_matched'] = 'Not a match'
                        break
            else:
                logger.warning("Color not found.")
                data['color_matched'] = 'Color not found.'
                data['color_found'] = False

    logger.debug("Back images comparison started.")
    for pc9, data in input_data.iteritems():
        product_loaded = data['product_loaded']
        color_found = data['color_found']
        if not product_loaded or not color_found:
            continue
        image_urls = data.get('image_urls')
        logger.debug('Processing product web_id={0}'.format(data.get('web_id')))

        if image_urls:
            for image_url in image_urls:
                if str(data['web_id']) in image_url and re.search(r'ALT\d*', image_url):
                    try:
                        logger.debug("Alt image found ({0}). Trying to compare..".format(image_url))
                        match_result = downloadS3.getComparison(pc9, seasons, 'B', image_url)
                    except Exception as e:
                        logger.error("Image matching error: {0}".format(e.message))
                        match_result = 0

                    logger.debug("Matching result = {0}".format(match_result))
                    if match_result > 98:
                        data['image_back'] = match_result
                        break
                    elif match_result > 95:
                        data['image_back'] = max(data.get('image_back', 0), match_result)

    logger.debug("Side images comparison started.")
    for pc9, data in input_data.iteritems():
        product_loaded = data['product_loaded']
        color_found = data['color_found']
        if not product_loaded or not color_found:
            continue
        image_urls = data.get('image_urls')
        logger.debug('Processing product web_id={0}'.format(data.get('web_id')))

        if image_urls:
            for image_url in image_urls:
                if str(data['web_id']) in image_url and re.search(r'ALT\d*', image_url):
                    try:
                        logger.debug("Alt image found ({0}). Trying to compare..".format(image_url))
                        match_result = downloadS3.getComparison(pc9, seasons, 'S', image_url)
                    except Exception as e:
                        logger.error("Image matching error: {0}".format(e.message))
                        match_result = 0

                    logger.debug("Matching result = {0}".format(match_result))
                    if match_result > 98:
                        data['image_side'] = match_result
                        break
                    elif match_result > 95:
                        data['image_side'] = max(data.get('image_side', 0), match_result)

    logger.debug("Writing results to output file {0}".format(output_file))
    workbook = xlsxwriter.Workbook(output_file)
    worksheet = workbook.add_worksheet()

    worksheet.write(0, 0, 'Web-ID')
    worksheet.write(0, 1, 'PC9 Tag')
    worksheet.write(0, 2, 'Color')
    worksheet.write(0, 3, 'Matched or Not')
    worksheet.write(0, 4, 'Alternate Images found')
    worksheet.write(0, 5, 'Video')
    worksheet.write(0, 6, 'Size Chart')

    row_index = 1

    for pc9, data in input_data.iteritems():
        worksheet.write(row_index, 0, data['web_id'])
        worksheet.write(row_index, 1, pc9)
        worksheet.write(row_index, 2, data['color'])
        worksheet.write(row_index, 3, data.get('color_matched') if data.get('color_matched') else '')

        if data.get('image_back') and data.get('image_side'):
            alternate_images_found = 'Side and back found'
        elif data.get('image_back'):
            alternate_images_found = 'Back found'
        elif data.get('image_side'):
            alternate_images_found = 'Side found'
        else:
            alternate_images_found = ''

        worksheet.write(row_index, 4, alternate_images_found)
        worksheet.write(row_index, 5, 'Yes' if data.get('video') else 'No')
        worksheet.write(row_index, 6, 'Yes' if data.get('size_chart') else 'No')

        row_index += 1

    workbook.close()
    logger.debug("Kohls products audit finished".format(output_file))
