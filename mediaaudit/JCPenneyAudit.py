import xlsxwriter
import JCPenneyScraperAux
import downloadS3
from openpyxl import load_workbook
import re


def get_column_index(regex,  columns):
    for index, column in enumerate(columns):
        if column and re.search(regex, column):
            return index


def run(input_file, output_file):
    # All data goes into here #
    input_data = {}
    seasons = ['F17', 'S17', 'F16', 'S16', 'F15', 'F14']

    # May need to re-name Sheet1 field #
    wb = load_workbook(filename=input_file, read_only=True)
    ws = wb['Sheet1']

    # Reading header
    columns = [cell.value.lower() if isinstance(cell.value, basestring) else cell.value for cell in ws[1]]

    web_id_index = get_column_index(r'web', columns)
    if web_id_index is None:
        web_id_index = 0

    pc9_index = get_column_index(r'pc9', columns)
    if pc9_index is None:
        pc9_index = 1

    color_index = get_column_index(r'wash|color', columns)
    if color_index is None:
        color_index = 2

    # Reading in from File Mapping
    for row in ws.iter_rows(min_row=2):
        pc9 = row[pc9_index].value

        if pc9:
            input_data[pc9] = {
                'web_id': row[web_id_index].value,
                'color': row[color_index].value
            }

    #  We use a set to avoid duplicate Product IDs #
    products_cache = {}

    for pc9, data in input_data.iteritems():
        web_id = data['web_id']

        if web_id not in products_cache:
            url = 'http://www.jcpenney.com/null/prod.jump?ppId=pp{}&catId=WebID'.format(web_id)
            products_cache[web_id] = JCPenneyScraperAux.load_product(url)

        product = products_cache[web_id]

        if product:
            data['color_url'] = JCPenneyScraperAux.get_color_url(product, data['color'])
            data['image_urls'] = JCPenneyScraperAux.get_image_urls(product)
            data['video'] = JCPenneyScraperAux.get_video(product)
            data['size_chart'] = JCPenneyScraperAux.get_size_chart(product)
            data['fit_guide'] = JCPenneyScraperAux.get_fit_guide(product)

    # Comparing fronts
    for pc9, data in input_data.iteritems():
        color_url = data.get('color_url')

        if color_url:
            try:
                match_result = downloadS3.getComparison(pc9, seasons, 'F', color_url)
            except:
                match_result = 0

            if match_result > 97:
                data['color_matched'] = 'Yes'
            elif match_result > 91:
                data['color_matched'] = 'Yes (cropped)'
            elif match_result > 90:
                data['color_matched'] = 'Indeterminate'
            else:
                data['color_matched'] = 'Not a match'

    # Comparing backs
    for pc9, data in input_data.iteritems():
        image_urls = data.get('image_urls')

        if image_urls:
            for image_url in image_urls:
                try:
                    match_result = downloadS3.getComparison(pc9, seasons, 'B', image_url)
                except:
                    match_result = 0

                if match_result > 98:
                    data['image_back'] = match_result
                    break
                elif match_result > 95:
                    data['image_back'] = max(data.get('image_back', 0), match_result)

    # Comparing sides
    for pc9, data in input_data.iteritems():
        image_urls = data.get('image_urls')

        if image_urls:
            for image_url in image_urls:
                try:
                    match_result = downloadS3.getComparison(pc9, seasons, 'S', image_url)
                except:
                    match_result = 0

                if match_result > 98:
                    data['image_side'] = match_result
                    break
                elif match_result > 95:
                    data['image_side'] = max(data.get('image_side', 0), match_result)

    # Writing out to the file
    workbook = xlsxwriter.Workbook(output_file)
    worksheet = workbook.add_worksheet()

    worksheet.write(0, 0, 'Web-ID')
    worksheet.write(0, 1, 'PC9 Tag')
    worksheet.write(0, 2, 'Color')
    worksheet.write(0, 3, 'Matched or Not')
    worksheet.write(0, 4, 'Alternate Images found')
    worksheet.write(0, 5, 'Video')
    worksheet.write(0, 6, 'Size Chart')
    worksheet.write(0, 7, 'Fit Guide')

    row_index = 1

    for pc9, data in input_data.iteritems():
        worksheet.write(row_index, 0, data['web_id'])
        worksheet.write(row_index, 1, pc9)
        worksheet.write(row_index, 2, data['color'])
        worksheet.write(row_index, 3, data.get('color_matched') if data.get('color_matched') else '')

        if data.get('image_back') and data.get('image_side'):
            alternate_images_found = 'Side and back found'
        elif data.get('image_back'):
            alternate_images_found = 'Back found'
        elif data.get('image_side'):
            alternate_images_found = 'Side found'
        else:
            alternate_images_found = ''

        worksheet.write(row_index, 4, alternate_images_found)
        worksheet.write(row_index, 5, 'Yes' if data.get('video') else 'No')
        worksheet.write(row_index, 6, 'Yes' if data.get('size_chart') else 'No')
        worksheet.write(row_index, 7, 'Yes' if data.get('fit_guide') else 'No')

        row_index += 1

    workbook.close()

