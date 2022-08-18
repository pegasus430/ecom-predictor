import xlsxwriter
import re
from openpyxl import load_workbook


def run(input_file, output_file):
    wb = load_workbook(filename=input_file, read_only=True)
    ws = wb['Sheet1']

    rows = []

    for row in ws.iter_rows(min_row=2, max_col=3):
        rows.append({
            'web_id': convert_web_id(row[0].value),
            'pc9': convert_pc9(row[1].value),
            'color': convert_color(row[2].value)
        })

    workbook = xlsxwriter.Workbook(output_file)
    worksheet = workbook.add_worksheet()

    worksheet.write(0, 0, 'Web-ID')
    worksheet.write(0, 1, 'PC9 Tag')
    worksheet.write(0, 2, 'Color')

    for row_index, row in enumerate(rows, start=1):
        worksheet.write(row_index, 0, row['web_id'])
        worksheet.write(row_index, 1, row['pc9'])
        worksheet.write(row_index, 2, row['color'])

    workbook.close()


def convert_web_id(web_id):
    if web_id is not None:
        web_id = re.sub(r'\D', '', str(web_id))[:10]

    return web_id


def convert_pc9(pc9):
    if pc9 is not None:
        pc9 = re.sub(r'\D', '', str(pc9))[:9]

    return pc9


def convert_color(color):
    if color is not None:
        color = color.strip()

    return color
